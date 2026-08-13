from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import random
import re
import shutil
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parents[1]
CATEGORY_PATH = PROJECT_ROOT / "data_seed" / "merged_categories.json"
UI_SCHEMA_PATH = PROJECT_ROOT / "data_seed" / "product_field_ui_schemas.json"
SCHEMA_VERSION = "batch_product_input_v4"
DEFAULT_TEST_USERNAME = "123@test"
SYNTHETIC_DATE = "2026-08-04"

CATEGORY_SUBCATEGORY_ALIASES = {
    ("消费电子", "智能手表"): ("消费电子", "智能手表/手环"),
    ("消费电子", "无线耳机"): ("消费电子", "真无线耳机"),
    ("消费电子", "游戏主机"): ("消费电子", "游戏机"),
    ("消费电子", "数码相机"): ("消费电子", "相机（含无人机）"),
    ("家用电器", "洗衣机"): ("家用电器", "洗衣机/烘干机"),
    ("家用电器", "微波炉"): ("家用电器", "微波炉/烤箱/蒸烤箱"),
    ("家用电器", "电饭煲"): ("家用电器", "电饭煲/压力锅"),
    ("个护健康", "血压计"): ("个护健康", "血压计/血糖仪"),
    ("个护健康", "血糖仪"): ("个护健康", "血压计/血糖仪"),
}

DEFAULT_SHOWCASE_CASE_IDS = {
    "SONY-PS5SLIM-01",
    "SONY-A7M4-01",
    "GREE-YHPRO-01",
    "mb_stroller_001",
    "case_portable_power_global",
    "case_electric_wheelchair_cn",
}

BRAND_ALIAS_GROUPS = (
    {"apple", "苹果"},
    {"huawei", "华为"},
    {"xiaomi", "小米"},
    {"sony", "索尼"},
    {"samsung", "三星"},
    {"lenovo", "联想"},
    {"microsoft", "微软"},
    {"nintendo", "任天堂"},
    {"canon", "佳能"},
    {"nikon", "尼康"},
    {"panasonic", "松下"},
)

SHEETS = {
    "instructions": "00_填写说明",
    "cases": "01_测试任务",
    "products": "02_产品主表",
    "params": "03_参数明细",
    "evidence": "04_证据来源",
    "dictionary": "05_品类参数字典",
    "profiles": "06_运行配置",
    "validation": "07_校验结果",
    "input_checks": "08_输入检查",
}

CASE_HEADERS = [
    "case_id", "enabled", "project_name", "target_product_id", "competitor_mode",
    "competitor_product_ids", "auto_competitor_count",
    "assumed_market_competitor_count", "notes",
]
PRODUCT_HEADERS = [
    "product_record_id", "product_name", "brand", "confirmed_sku", "category", "subcategory",
    "variant_description", "price_cny", "price_type", "price_status", "sales_channel", "product_url",
    "release_date", "collection_date", "collector", "review_status", "notes",
]
PARAM_HEADERS = [
    "product_record_id", "field_code", "field_name_cn", "field_value", "value_type", "unit",
    "data_status", "collection_date", "notes",
]
EVIDENCE_HEADERS = [
    "evidence_id", "product_record_id", "field_code", "source_type", "source_name", "source_url",
    "source_title", "observed_value", "observed_date", "collection_date", "confidence",
    "requires_manual_review", "source_summary",
]
DICTIONARY_HEADERS = [
    "category", "subcategory", "field_code", "field_name_cn", "field_type", "description", "unit",
    "required", "control_type", "min", "max", "options", "hint", "default_weight",
]
PROFILE_HEADERS = [
    "case_id", "target_crowd", "crowd_profile", "crowd_segments", "strategies",
    "strategy_details", "scenes", "scene_details", "sample_size", "decision_weight_profile",
    "social_propagation_config",
]
VALIDATION_HEADERS = [
    "case_id", "target_product_id", "competitor_product_ids", "field_code", "field_name_cn",
    "target_value", "competitor_values", "comparison_method", "normalized_gap", "default_weight",
    "data_completeness_pct", "input_fit_score", "fit_interpretation",
]
INPUT_CHECK_HEADERS = [
    "case_id", "product_record_id", "sheet", "row_number", "field", "severity", "error_code",
    "message", "suggested_fix",
]

FILL_REQUIRED = PatternFill("solid", fgColor="FFF2CC")
FILL_CONDITIONAL = PatternFill("solid", fgColor="FFF9E6")
FILL_SYSTEM = PatternFill("solid", fgColor="E7E6E6")
FILL_PLATFORM = PatternFill("solid", fgColor="DDEBF7")
FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FONT_HEADER = Font(color="FFFFFF", bold=True)


@dataclass
class ValidationIssue:
    case_id: str = ""
    product_record_id: str = ""
    sheet: str = ""
    row_number: int | str = ""
    field: str = ""
    severity: str = "error"
    error_code: str = ""
    message: str = ""
    suggested_fix: str = ""

    def as_row(self) -> list[Any]:
        return [getattr(self, key) for key in INPUT_CHECK_HEADERS]


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def canonical_category_subcategory(category: Any, subcategory: Any) -> tuple[str, str]:
    value = (clean_text(category), clean_text(subcategory))
    return CATEGORY_SUBCATEGORY_ALIASES.get(value, value)


def canonical_brand(value: Any) -> str:
    brand = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", clean_text(value).lower())
    for index, aliases in enumerate(BRAND_ALIAS_GROUPS):
        if brand in aliases:
            return f"brand_group_{index}"
    return brand


def product_model_signature(product: dict[str, Any]) -> str:
    text = clean_text(product.get("product_name")).lower()
    brand = canonical_brand(product.get("brand"))
    aliases: set[str] = set()
    if brand.startswith("brand_group_"):
        aliases = BRAND_ALIAS_GROUPS[int(brand.rsplit("_", 1)[1])]
    for alias in sorted(aliases, key=len, reverse=True):
        text = text.replace(alias, "")
    for token in ("轻薄版", "slim", "英寸", "版本", "版", "款"):
        text = text.replace(token, "")
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)


def is_same_library_product(target: dict[str, Any], candidate: dict[str, Any]) -> bool:
    if canonical_brand(target.get("brand")) != canonical_brand(candidate.get("brand")):
        return False
    target_signature = product_model_signature(target)
    candidate_signature = product_model_signature(candidate)
    if len(target_signature) < 4 or target_signature != candidate_signature:
        return False
    target_price = safe_number(target.get("price_cny"))
    candidate_price = safe_number(candidate.get("price_cny"))
    if target_price is None or candidate_price is None or target_price <= 0 or candidate_price <= 0:
        return True
    return abs(candidate_price - target_price) / max(target_price, candidate_price) <= 0.25


def split_values(value: Any) -> list[str]:
    return [item.strip() for item in clean_text(value).replace("，", "|").split("|") if item.strip()]


def truthy(value: Any) -> bool:
    return clean_text(value).lower() in {"1", "true", "yes", "y", "是", "启用"}


def is_showcase_case(case: dict[str, Any]) -> bool:
    return bool(
        clean_text(case.get("case_id")) in DEFAULT_SHOWCASE_CASE_IDS
        or clean_text(case.get("project_name")).startswith("【代表案例】")
        or "代表案例" in clean_text(case.get("notes"))
        or truthy(case.get("showcase"))
    )


def safe_number(value: Any) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def has_product_price_provenance(product: dict[str, Any]) -> bool:
    """Accept structured product-row provenance when no duplicate evidence row exists."""
    url = clean_text(product.get("product_url"))
    return bool(clean_text(product.get("price_type")) and url.startswith(("http://", "https://")))


def json_value(value: Any, fallback: Any) -> Any:
    if value in (None, ""):
        return fallback
    if isinstance(value, (dict, list)):
        return value
    return json.loads(str(value))


def load_category_dictionary() -> list[dict[str, Any]]:
    payload = json.loads(CATEGORY_PATH.read_text(encoding="utf-8"))
    ui_payload = json.loads(UI_SCHEMA_PATH.read_text(encoding="utf-8")) if UI_SCHEMA_PATH.exists() else {"rules": []}
    ui_map: dict[tuple[str, str, str], dict[str, Any]] = {}
    for rule in ui_payload.get("rules", []):
        category = clean_text(rule.get("category"))
        subcategory = clean_text(rule.get("subcategory"))
        for field_code, schema in (rule.get("fields") or {}).items():
            if isinstance(schema, dict):
                ui_map[(category, subcategory, clean_text(field_code))] = schema

    rows: list[dict[str, Any]] = []
    for category in payload.get("categories", []):
        category_name = clean_text(category.get("category"))
        subcategory = clean_text(category.get("subcategory"))
        for field in category.get("fields", []):
            code = clean_text(field.get("name"))
            if code == "price_cny":
                continue
            schema = ui_map.get((category_name, subcategory, code), {})
            rows.append(
                {
                    "category": category_name,
                    "subcategory": subcategory,
                    "field_code": code,
                    "field_name_cn": clean_text(schema.get("label") or field.get("desc") or code),
                    "field_type": clean_text(field.get("type") or "string"),
                    "description": clean_text(field.get("desc")),
                    "unit": clean_text(schema.get("unit")),
                    "required": "否",
                    "control_type": clean_text(schema.get("controlType")),
                    "min": schema.get("min", ""),
                    "max": schema.get("max", ""),
                    "options": "|".join(str(item) for item in (schema.get("options") or [])),
                    "hint": clean_text(schema.get("hint") or ui_payload.get("default_hint")),
                    "default_weight": float(schema.get("defaultWeight") or 1.0),
                }
            )
    return rows


def rows_from_sheet(workbook: Any, name: str) -> list[dict[str, Any]]:
    if name not in workbook.sheetnames:
        return []
    sheet = workbook[name]
    headers = [clean_text(cell.value) for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not any(value not in (None, "") for value in values):
            continue
        item = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        item["_row_number"] = row_number
        rows.append(item)
    return rows


def setup_sheet(sheet: Any, headers: list[str], *, platform: bool = False) -> None:
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).coordinate}"
    for cell in sheet[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, header in enumerate(headers, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(42, max(14, len(header) + 4))
        if platform:
            sheet.cell(1, index).fill = FILL_PLATFORM


def add_list_validation(sheet: Any, column: str, values: list[str], max_row: int = 5000) -> None:
    validation = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{column}2:{column}{max_row}")


def market_profiles() -> list[dict[str, Any]]:
    common_weights = {"template": "default"}
    profiles = [
        {
            "market_profile_key": "default",
            "target_crowd": "目标品类潜在消费者",
            "crowd_profile": {"price_sensitivity": "medium", "feature_priorities": ["核心功能", "可靠性", "价格"]},
            "strategies": ["卖点强化", "内容种草"],
            "scenes": ["综合电商"],
        },
        {
            "market_profile_key": "personal_care",
            "target_crowd": "关注日常护理与性价比的年轻消费者",
            "crowd_profile": {
                "age_range": "22-40", "city_tier": "一线/二线", "income_level": "中等收入",
                "life_stage": "年轻白领与家庭用户", "price_sensitivity": "medium",
                "feature_priorities": ["清洁效果", "续航", "使用舒适度"],
                "channel_preferences": ["综合电商", "内容种草"],
                "purchase_motivations": ["护理升级", "替换旧设备"], "risk_concerns": ["效果真实性", "耗材成本"],
                "custom_description": "模拟个护健康市场配置",
            },
            "strategies": ["测评背书", "卖点强化", "内容种草"],
            "scenes": ["综合电商", "日常家庭护理"],
        },
        {
            "market_profile_key": "consumer_electronics",
            "target_crowd": "关注健康与智能体验的数码消费者",
            "crowd_profile": {
                "age_range": "20-45", "city_tier": "一线/新一线", "income_level": "中高收入",
                "life_stage": "科技尝鲜与运动健康人群", "price_sensitivity": "medium",
                "feature_priorities": ["健康监测", "续航", "系统兼容"],
                "channel_preferences": ["品牌旗舰店", "科技媒体", "综合电商"],
                "purchase_motivations": ["健康管理", "体验升级"], "risk_concerns": ["数据准确性", "续航衰减"],
                "custom_description": "模拟消费电子市场配置",
            },
            "strategies": ["测评背书", "场景包装", "内容种草"],
            "scenes": ["线上首发", "运动健康"],
        },
    ]
    result: list[dict[str, Any]] = []
    for profile in profiles:
        target = profile["target_crowd"]
        crowd_profile = profile["crowd_profile"]
        strategies = profile["strategies"]
        scenes = profile["scenes"]
        result.append(
            {
                **profile,
                "case_id": profile["market_profile_key"],
                "crowd_segments": [{"name": target, "ratio": 100, "is_custom": True, "profile": crowd_profile}],
                "strategy_details": {name: {"actions": [f"围绕{name}执行模拟方案"], "channels": ["综合电商"]} for name in strategies},
                "scene_details": {name: {"place": name, "frequency": "每周", "purchase_trigger": "需求升级", "pain_point": "信息不充分", "decision_maker": "本人", "scene_tags": [name], "note": "平台预置模拟场景"} for name in scenes},
                "sample_size": 10000,
                "decision_weight_profile": common_weights,
                "social_propagation_config": {"external_traffic_per_round": 100, "scene_fission_factor": 1.1},
            }
        )
    return result


def create_workbook(*, include_samples: bool = False, sample_count: int = 2, seed: int = 20260804) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    instructions = workbook.create_sheet(SHEETS["instructions"])
    instructions.append(["AgentSim批量产品信息采集模板", SCHEMA_VERSION])
    instructions.append(["用途", "外部采集产品事实，并按case_id自由设置目标人群、市场策略、场景和仿真配置。"])
    instructions.append(["竞品规则", "competitor_product_ids留空时全部自动检索；填写时将其作为自定义竞品，并从产品库自动补足auto_competitor_count。competitor_mode为旧版兼容字段，可以留空。"])
    instructions.append(["专用测试账号", DEFAULT_TEST_USERNAME])
    instructions.append(["公开演示初始密码", "123456"])
    instructions.append(["账号信息位置", "上方仅展示已明确公开的演示凭据；本地通过AGENTSIM_API_BASE / AGENTSIM_TEST_USERNAME / AGENTSIM_TEST_PASSWORD环境变量登录。其他账号密码和Token不得写入Excel。"])
    instructions.append(["列表格式", "多个值使用 | 分隔"])
    instructions.append(["运行配置格式", "06表每个case_id填写一行；strategies/scenes为JSON数组，crowd_profile/strategy_details/scene_details/权重/传播配置为JSON对象。策略内容由交付方自由设置。"])
    instructions.append(["数据状态", "confirmed / estimated / manual / missing"])
    instructions.append(["校验结果口径", "07表仅比较价格与已输入关键参数，input_fit_score是目标/竞品输入可比程度，不是市场预测准确率；格式错误见08表。"])
    instructions.append(["模拟数据", "所有模拟记录均为synthetic，不代表真实产品或市场事实。"])
    instructions["Z1"] = SCHEMA_VERSION
    instructions.column_dimensions["Z"].hidden = True
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 100
    instructions.sheet_state = "visible"

    cases = workbook.create_sheet(SHEETS["cases"])
    setup_sheet(cases, CASE_HEADERS)
    products = workbook.create_sheet(SHEETS["products"])
    setup_sheet(products, PRODUCT_HEADERS)
    params = workbook.create_sheet(SHEETS["params"])
    setup_sheet(params, PARAM_HEADERS)
    evidence = workbook.create_sheet(SHEETS["evidence"])
    setup_sheet(evidence, EVIDENCE_HEADERS)
    dictionary = workbook.create_sheet(SHEETS["dictionary"])
    setup_sheet(dictionary, DICTIONARY_HEADERS, platform=True)
    profiles = workbook.create_sheet(SHEETS["profiles"])
    setup_sheet(profiles, PROFILE_HEADERS, platform=True)
    validation = workbook.create_sheet(SHEETS["validation"])
    setup_sheet(validation, VALIDATION_HEADERS, platform=True)
    input_checks = workbook.create_sheet(SHEETS["input_checks"])
    setup_sheet(input_checks, INPUT_CHECK_HEADERS, platform=True)

    for row in load_category_dictionary():
        dictionary.append([row.get(header, "") for header in DICTIONARY_HEADERS])
    dictionary.protection.sheet = True

    # 运行配置由交付方按任务自由填写；空白模板不注入市场策略。

    add_list_validation(cases, "B", ["是", "否"])
    add_list_validation(cases, "E", ["excel", "library_auto"])
    add_list_validation(products, "I", ["日常价", "活动价", "官方指导价", "历史均价"])
    add_list_validation(products, "J", ["confirmed", "estimated", "manual", "missing"])
    add_list_validation(products, "P", ["pending", "reviewed", "rejected"])
    add_list_validation(params, "E", ["string", "number", "boolean"])
    add_list_validation(params, "G", ["confirmed", "estimated", "manual", "missing"])
    add_list_validation(evidence, "D", ["official", "ecommerce", "media", "manual", "other"])
    add_list_validation(evidence, "L", ["是", "否"])

    required_columns = {
        SHEETS["cases"]: {1, 2, 3, 4, 8},
        SHEETS["products"]: {1, 2, 5, 6, 8, 9, 10, 14, 15, 16},
        SHEETS["params"]: {1, 2, 3, 4, 5, 7, 8},
        SHEETS["evidence"]: {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13},
        SHEETS["profiles"]: {1, 2, 3, 5, 7, 9},
    }
    for sheet_name, columns in required_columns.items():
        sheet = workbook[sheet_name]
        for column in columns:
            sheet.cell(1, column).fill = FILL_REQUIRED
            sheet.cell(1, column).font = Font(color="000000", bold=True)

    if include_samples:
        populate_samples(workbook, sample_count=sample_count, seed=seed)
    return workbook


def field_rows_for(category: str, subcategory: str) -> list[dict[str, Any]]:
    return [row for row in load_category_dictionary() if row["category"] == category and row["subcategory"] == subcategory]


def populate_samples(workbook: Workbook, *, sample_count: int, seed: int) -> None:
    rng = random.Random(seed)
    definitions = [
        {
            "case_id": "demo_toothbrush_001", "project_name": "模拟电动牙刷批量测试",
            "profile": "personal_care", "category": "个护健康", "subcategory": "电动牙刷",
            "products": [
                ("DEMO-TB-A1", "智测声波牙刷A1", "模拟品牌A", "SYN-TB-A1", 399),
                ("DEMO-TB-B1", "智测声波牙刷B1", "模拟品牌B", "SYN-TB-B1", 299),
                ("DEMO-TB-C1", "智测声波牙刷C1", "模拟品牌C", "SYN-TB-C1", 599),
            ],
            "values": {
                "vibration_frequency": [31000, 28000, 40000], "cleaning_modes": [5, 3, 7],
                "battery_life": [30, 20, 45], "charging_method": ["USB-C", "感应充电", "无线充电杯"],
                "waterproof_rating": ["IPX7", "IPX6", "IPX8"], "pressure_sensor": ["是", "否", "是"],
                "brush_head_type": ["柔软清洁型", "基础清洁型", "多功能护理型"],
            },
        },
        {
            "case_id": "demo_watch_001", "project_name": "模拟智能手表批量测试",
            "profile": "consumer_electronics", "category": "消费电子", "subcategory": "智能手表/手环",
            "products": [
                ("DEMO-WATCH-A1", "智测健康手表A1", "模拟品牌A", "SYN-WATCH-A1", 1299),
                ("DEMO-WATCH-B1", "智测运动手表B1", "模拟品牌B", "SYN-WATCH-B1", 899),
                ("DEMO-WATCH-C1", "智测高端手表C1", "模拟品牌C", "SYN-WATCH-C1", 1799),
            ],
            "values": {
                "screen": ["1.75英寸AMOLED", "1.6英寸LCD", "1.9英寸LTPO AMOLED"],
                "health_monitoring": ["心率|血氧|睡眠", "心率|睡眠", "心率|血氧|睡眠|体温趋势"],
                "sports_modes": [120, 80, 150], "battery_life": [12, 18, 10],
                "gps": ["独立双频GPS", "独立GPS", "独立双频多星GPS"],
                "waterproof_rating": ["5ATM", "IP68", "10ATM"],
                "compatible_os": ["Android|iOS", "Android|iOS", "Android|iOS"],
            },
        },
    ]
    cases_sheet = workbook[SHEETS["cases"]]
    products_sheet = workbook[SHEETS["products"]]
    params_sheet = workbook[SHEETS["params"]]
    evidence_sheet = workbook[SHEETS["evidence"]]
    profiles_sheet = workbook[SHEETS["profiles"]]
    sample_profiles = {profile["market_profile_key"]: profile for profile in market_profiles()}

    for definition in definitions[: max(1, min(sample_count, 2))]:
        product_ids = [item[0] for item in definition["products"]]
        cases_sheet.append(
            [definition["case_id"], "是", definition["project_name"], product_ids[0], "excel",
             "|".join(product_ids[1:]), 2, 20, "synthetic fixture；仅用于测试"]
        )
        profile = sample_profiles[definition["profile"]]
        profiles_sheet.append(
            [
                definition["case_id"], profile["target_crowd"],
                json.dumps(profile["crowd_profile"], ensure_ascii=False),
                json.dumps(profile["crowd_segments"], ensure_ascii=False),
                json.dumps(profile["strategies"], ensure_ascii=False),
                json.dumps(profile["strategy_details"], ensure_ascii=False),
                json.dumps(profile["scenes"], ensure_ascii=False),
                json.dumps(profile["scene_details"], ensure_ascii=False),
                profile["sample_size"], json.dumps(profile["decision_weight_profile"], ensure_ascii=False),
                json.dumps(profile["social_propagation_config"], ensure_ascii=False),
            ]
        )
        dictionary_rows = {row["field_code"]: row for row in field_rows_for(definition["category"], definition["subcategory"])}
        for product_index, (record_id, name, brand, sku, price) in enumerate(definition["products"]):
            products_sheet.append(
                [record_id, name, brand, sku, definition["category"], definition["subcategory"],
                 f"模拟完整配置{product_index + 1}", price, "日常价", "manual", "模拟综合电商",
                 f"https://example.invalid/agentsim/{record_id}", "2026-01-01", SYNTHETIC_DATE,
                 "AgentSim Synthetic Fixture", "reviewed", "synthetic；不代表真实产品"]
            )
            evidence_values = {"price_cny": price}
            for code, values in definition["values"].items():
                dictionary_row = dictionary_rows[code]
                value = values[product_index]
                params_sheet.append(
                    [record_id, code, dictionary_row["field_name_cn"], value, dictionary_row["field_type"],
                     dictionary_row["unit"] or "无", "manual", SYNTHETIC_DATE, "synthetic full-field fixture"]
                )
                evidence_values[code] = value
            for evidence_index, (field_code, value) in enumerate(evidence_values.items(), 1):
                evidence_sheet.append(
                    [f"E-{record_id}-{evidence_index:02d}", record_id, field_code, "manual",
                     "AgentSim Synthetic Fixture", f"https://example.invalid/agentsim/{record_id}/{field_code}",
                     "AgentSim模拟数据字段说明", str(value), SYNTHETIC_DATE, SYNTHETIC_DATE, 1.0, "是",
                     f"synthetic；仅用于验证字段 {field_code}，不代表真实商品资料；随机标记{rng.randint(1000, 9999)}"]
                )


def write_workbook(workbook: Workbook, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def dictionary_map(workbook: Any) -> dict[tuple[str, str, str], dict[str, Any]]:
    result: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows_from_sheet(workbook, SHEETS["dictionary"]):
        key = (clean_text(row.get("category")), clean_text(row.get("subcategory")), clean_text(row.get("field_code")))
        result[key] = row
    return result


def validate_workbook(path: Path) -> tuple[list[ValidationIssue], dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    issues: list[ValidationIssue] = []
    for sheet_name in SHEETS.values():
        if sheet_name not in workbook.sheetnames:
            issues.append(ValidationIssue(sheet=sheet_name, error_code="SHEET_MISSING", message="缺少工作表", suggested_fix="使用官方模板重新填写"))
    if issues:
        return issues, {}

    cases = rows_from_sheet(workbook, SHEETS["cases"])
    products = rows_from_sheet(workbook, SHEETS["products"])
    params = rows_from_sheet(workbook, SHEETS["params"])
    evidence = rows_from_sheet(workbook, SHEETS["evidence"])
    profiles = rows_from_sheet(workbook, SHEETS["profiles"])
    field_map = dictionary_map(workbook)

    def issue(*, row: dict[str, Any] | None = None, **kwargs: Any) -> None:
        if row is not None:
            kwargs.setdefault("row_number", row.get("_row_number", ""))
        issues.append(ValidationIssue(**kwargs))

    case_ids: set[str] = set()
    product_ids: set[str] = set()
    profile_rows: dict[str, dict[str, Any]] = {}
    for profile in profiles:
        config_case_id = clean_text(profile.get("case_id"))
        if not config_case_id:
            issue(row=profile, sheet=SHEETS["profiles"], field="case_id", error_code="CONFIG_CASE_REQUIRED", message="运行配置必须关联case_id", suggested_fix="填写测试任务中的case_id")
            continue
        if config_case_id in profile_rows:
            issue(row=profile, case_id=config_case_id, sheet=SHEETS["profiles"], field="case_id", error_code="CONFIG_CASE_DUPLICATE", message="同一case_id存在多条运行配置", suggested_fix="每个任务仅保留一条配置")
        profile_rows[config_case_id] = profile
        for field in ("target_crowd", "strategies", "scenes"):
            if clean_text(profile.get(field)) == "":
                issue(row=profile, case_id=config_case_id, sheet=SHEETS["profiles"], field=field, error_code="CONFIG_FIELD_REQUIRED", message=f"运行配置字段 {field} 不能为空", suggested_fix="按任务自由填写")
        for field, expected_type in (("crowd_profile", dict), ("crowd_segments", list), ("strategies", list), ("strategy_details", dict), ("scenes", list), ("scene_details", dict), ("decision_weight_profile", dict), ("social_propagation_config", dict)):
            value = profile.get(field)
            if value in (None, "") and field in {"crowd_segments", "strategy_details", "scene_details", "decision_weight_profile", "social_propagation_config"}:
                continue
            try:
                parsed = json_value(value, expected_type())
            except (json.JSONDecodeError, TypeError):
                issue(row=profile, case_id=config_case_id, sheet=SHEETS["profiles"], field=field, error_code="CONFIG_JSON_INVALID", message=f"{field}不是有效JSON", suggested_fix="按填写说明提供JSON数组或对象")
                continue
            if not isinstance(parsed, expected_type):
                issue(row=profile, case_id=config_case_id, sheet=SHEETS["profiles"], field=field, error_code="CONFIG_JSON_TYPE", message=f"{field}的数据类型不正确", suggested_fix=f"填写JSON {expected_type.__name__}")
    products_by_id: dict[str, dict[str, Any]] = {}
    duplicate_identity: dict[tuple[str, str, str], str] = {}

    for row in products:
        record_id = clean_text(row.get("product_record_id"))
        if not record_id:
            issue(row=row, sheet=SHEETS["products"], field="product_record_id", error_code="PRODUCT_ID_REQUIRED", message="产品记录编号不能为空", suggested_fix="填写唯一编号")
            continue
        if record_id in product_ids:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["products"], field="product_record_id", error_code="PRODUCT_ID_DUPLICATE", message="产品记录编号重复", suggested_fix="修改为唯一编号")
        product_ids.add(record_id)
        original_category = clean_text(row.get("category"))
        original_subcategory = clean_text(row.get("subcategory"))
        canonical_category, canonical_subcategory = canonical_category_subcategory(original_category, original_subcategory)
        if (canonical_category, canonical_subcategory) != (original_category, original_subcategory):
            row["category"] = canonical_category
            row["subcategory"] = canonical_subcategory
            issue(
                row=row,
                product_record_id=record_id,
                sheet=SHEETS["products"],
                field="subcategory",
                severity="info",
                error_code="CATEGORY_ALIAS_NORMALIZED",
                message=f"平台已自动将 {original_category}/{original_subcategory} 归一化为 {canonical_category}/{canonical_subcategory}",
                suggested_fix="无需重新调查竞品；建议后续源表直接使用平台标准小类",
            )
        products_by_id[record_id] = row
        for field in ("product_name", "category", "subcategory", "price_type", "price_status", "collection_date", "collector", "review_status"):
            if clean_text(row.get(field)) == "":
                issue(row=row, product_record_id=record_id, sheet=SHEETS["products"], field=field, error_code="PRODUCT_FIELD_REQUIRED", message=f"产品字段 {field} 不能为空", suggested_fix="补充产品事实")
        price = safe_number(row.get("price_cny"))
        status = clean_text(row.get("price_status"))
        if status == "missing" and price is not None:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["products"], field="price_cny", error_code="PRICE_STATUS_CONFLICT", message="价格状态为missing时价格应为空", suggested_fix="清空价格或修改状态")
        elif status != "missing" and (price is None or price <= 0):
            issue(row=row, product_record_id=record_id, sheet=SHEETS["products"], field="price_cny", error_code="PRICE_INVALID", message="非missing价格必须为正数", suggested_fix="填写大于0的人民币价格")
        identity = (clean_text(row.get("brand")).lower(), clean_text(row.get("confirmed_sku")).lower(), clean_text(row.get("variant_description")).lower())
        if identity[0] and identity[1] and identity in duplicate_identity:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["products"], field="confirmed_sku", error_code="PRODUCT_IDENTITY_DUPLICATE", message=f"与 {duplicate_identity[identity]} 品牌/SKU/版本重复", suggested_fix="合并重复产品记录")
        duplicate_identity[identity] = record_id

    params_by_product: dict[str, list[dict[str, Any]]] = defaultdict(list)
    seen_params: set[tuple[str, str]] = set()
    for row in params:
        record_id = clean_text(row.get("product_record_id"))
        code = clean_text(row.get("field_code"))
        if record_id not in products_by_id:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["params"], field="product_record_id", error_code="PARAM_PRODUCT_UNKNOWN", message="参数引用的产品不存在", suggested_fix="检查产品记录编号")
            continue
        if not code:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["params"], field="field_code", error_code="PARAM_CODE_REQUIRED", message="参数编码不能为空", suggested_fix="从参数字典选择编码")
            continue
        key = (record_id, code)
        if key in seen_params:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["params"], field="field_code", error_code="PARAM_DUPLICATE", message="同一产品参数编码重复", suggested_fix="保留一条参数记录")
        seen_params.add(key)
        params_by_product[record_id].append(row)
        product = products_by_id[record_id]
        dictionary_key = (clean_text(product.get("category")), clean_text(product.get("subcategory")), code)
        if dictionary_key not in field_map:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["params"], field="field_code", severity="warning", error_code="CUSTOM_PARAM", message="参数不在当前品类字典中，将作为自定义参数", suggested_fix="确认编码、名称、类型和单位")
        if row.get("field_value") in (None, "") and clean_text(row.get("data_status")) != "missing":
            issue(row=row, product_record_id=record_id, sheet=SHEETS["params"], field="field_value", error_code="PARAM_VALUE_REQUIRED", message="参数值为空但状态不是missing", suggested_fix="填写值或修改数据状态")
        if clean_text(row.get("value_type")) == "number" and row.get("field_value") not in (None, "") and safe_number(row.get("field_value")) is None:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["params"], field="field_value", error_code="PARAM_NUMBER_INVALID", message="数字参数无法解析", suggested_fix="只填写数字，单位单独填写")

    evidence_by_product_field: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    evidence_ids: set[str] = set()
    for row in evidence:
        evidence_id = clean_text(row.get("evidence_id"))
        record_id = clean_text(row.get("product_record_id"))
        field_code = clean_text(row.get("field_code"))
        if not evidence_id or evidence_id in evidence_ids:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["evidence"], field="evidence_id", error_code="EVIDENCE_ID_INVALID", message="证据编号为空或重复", suggested_fix="填写唯一证据编号")
        evidence_ids.add(evidence_id)
        if record_id not in products_by_id:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["evidence"], field="product_record_id", error_code="EVIDENCE_PRODUCT_UNKNOWN", message="证据引用的产品不存在", suggested_fix="检查产品记录编号")
        evidence_by_product_field[(record_id, field_code)].append(row)
        confidence = safe_number(row.get("confidence"))
        if confidence is None or not 0 <= confidence <= 1:
            issue(row=row, product_record_id=record_id, sheet=SHEETS["evidence"], field="confidence", error_code="CONFIDENCE_INVALID", message="置信度必须在0到1之间", suggested_fix="填写0～1的小数")
        url = clean_text(row.get("source_url"))
        if not url.startswith(("http://", "https://")):
            issue(row=row, product_record_id=record_id, sheet=SHEETS["evidence"], field="source_url", error_code="SOURCE_URL_INVALID", message="来源URL格式无效", suggested_fix="填写http或https URL")

    enabled_cases: list[dict[str, Any]] = []
    target_ids: set[str] = set()
    for row in cases:
        if not truthy(row.get("enabled")):
            continue
        enabled_cases.append(row)
        case_id = clean_text(row.get("case_id"))
        target_id = clean_text(row.get("target_product_id"))
        if not re.fullmatch(r"[A-Za-z0-9_-]+", case_id or ""):
            issue(row=row, case_id=case_id, sheet=SHEETS["cases"], field="case_id", error_code="CASE_ID_INVALID", message="case_id只能包含字母、数字、下划线和短横线", suggested_fix="修改任务编号")
        if case_id in case_ids:
            issue(row=row, case_id=case_id, sheet=SHEETS["cases"], field="case_id", error_code="CASE_ID_DUPLICATE", message="case_id重复", suggested_fix="修改为唯一编号")
        case_ids.add(case_id)
        target_ids.add(target_id)
        if target_id not in products_by_id:
            issue(row=row, case_id=case_id, product_record_id=target_id, sheet=SHEETS["cases"], field="target_product_id", error_code="TARGET_UNKNOWN", message="目标产品不存在", suggested_fix="检查产品记录编号")
        else:
            price = safe_number(products_by_id[target_id].get("price_cny"))
            if price is None or price <= 0:
                issue(row=row, case_id=case_id, product_record_id=target_id, sheet=SHEETS["cases"], field="target_product_id", error_code="TARGET_PRICE_REQUIRED", message="目标产品必须有确定价格", suggested_fix="补充目标产品价格")
            if not evidence_by_product_field.get((target_id, "price_cny")) and not has_product_price_provenance(products_by_id[target_id]):
                issue(row=row, case_id=case_id, product_record_id=target_id, sheet=SHEETS["cases"], field="target_product_id", error_code="TARGET_PRICE_EVIDENCE_REQUIRED", message="目标产品价格缺少证据", suggested_fix="在证据来源表补充price_cny证据")
        competitor_ids = split_values(row.get("competitor_product_ids"))
        if len(competitor_ids) != len(set(competitor_ids)):
            issue(row=row, case_id=case_id, sheet=SHEETS["cases"], field="competitor_product_ids", error_code="COMPETITOR_DUPLICATE", message="竞品编号重复", suggested_fix="去除重复编号")
        for competitor_id in competitor_ids:
            if competitor_id == target_id:
                issue(row=row, case_id=case_id, product_record_id=target_id, sheet=SHEETS["cases"], field="competitor_product_ids", error_code="SELF_COMPETITOR", message="目标产品不能作为自己的竞品", suggested_fix="更换竞品")
            elif competitor_id not in products_by_id:
                issue(row=row, case_id=case_id, product_record_id=competitor_id, sheet=SHEETS["cases"], field="competitor_product_ids", error_code="COMPETITOR_UNKNOWN", message="竞品编号不存在", suggested_fix="检查产品记录编号")
        if case_id not in profile_rows:
            issue(row=row, case_id=case_id, sheet=SHEETS["profiles"], field="case_id", error_code="CONFIG_UNKNOWN", message="当前任务缺少自由运行配置", suggested_fix="在06_运行配置新增同case_id的一行")
        auto_count = safe_number(row.get("auto_competitor_count"))
        if auto_count is not None and not 1 <= auto_count <= 50:
            issue(row=row, case_id=case_id, sheet=SHEETS["cases"], field="auto_competitor_count", error_code="AUTO_COMPETITOR_COUNT_INVALID", message="竞品目标数量必须为1～50", suggested_fix="留空使用默认值5，或填写1～50")
        assumed_count = safe_number(row.get("assumed_market_competitor_count"))
        if assumed_count is not None and not 5 <= assumed_count <= 50:
            issue(row=row, case_id=case_id, sheet=SHEETS["cases"], field="assumed_market_competitor_count", error_code="MARKET_COUNT_INVALID", message="市场竞品数量必须为5～50", suggested_fix="留空使用默认值20，或填写5～50")

    for target_id in target_ids:
        if target_id and not params_by_product.get(target_id):
            issue(product_record_id=target_id, sheet=SHEETS["params"], field="product_record_id", severity="warning", error_code="TARGET_PARAMS_EMPTY", message="目标产品没有参数明细", suggested_fix="补充核心参数")

    context = {
        "workbook": workbook, "cases": enabled_cases, "products": products_by_id,
        "params": params_by_product, "evidence": evidence_by_product_field,
        "profiles": profile_rows, "dictionary": field_map,
    }
    return issues, context


def custom_competitor_ids(case: dict[str, Any], context: dict[str, Any]) -> list[str]:
    """Use explicit IDs first, otherwise infer the documented TARGET/Cxx family convention."""
    explicit = split_values(case.get("competitor_product_ids"))
    if explicit:
        return explicit
    target_id = clean_text(case.get("target_product_id"))
    match = re.fullmatch(r"(.+)-\d+", target_id)
    if not match:
        return []
    prefix = match.group(1)
    target = context.get("products", {}).get(target_id, {})
    category = clean_text(target.get("category"))
    subcategory = clean_text(target.get("subcategory"))
    pattern = re.compile(rf"{re.escape(prefix)}-C\d+", re.IGNORECASE)
    return sorted(
        product_id
        for product_id, product in context.get("products", {}).items()
        if product_id != target_id
        and pattern.fullmatch(product_id)
        and clean_text(product.get("category")) == category
        and clean_text(product.get("subcategory")) == subcategory
    )


def build_input_fit_rows(context: dict[str, Any]) -> list[list[Any]]:
    """Build an input-comparability score, not a real-market accuracy score."""
    rows: list[list[Any]] = []
    if not context:
        return rows
    for case in context.get("cases", []):
        case_id = clean_text(case.get("case_id"))
        target_id = clean_text(case.get("target_product_id"))
        competitor_ids = custom_competitor_ids(case, context)
        if target_id not in context.get("products", {}):
            continue
        target_product = context["products"][target_id]
        target_params = {clean_text(item.get("field_code")): item for item in context["params"].get(target_id, [])}
        fields: list[tuple[str, str, Any, float, str]] = [
            ("price_cny", "价格", target_product.get("price_cny"), 1.5, "numeric_relative_gap")
        ]
        category = clean_text(target_product.get("category"))
        subcategory = clean_text(target_product.get("subcategory"))
        for code, item in target_params.items():
            dictionary_row = context["dictionary"].get((category, subcategory, code), {})
            fields.append(
                (
                    code,
                    clean_text(item.get("field_name_cn") or dictionary_row.get("field_name_cn") or code),
                    item.get("field_value"),
                    safe_number(dictionary_row.get("default_weight")) or 1.0,
                    "numeric_relative_gap" if clean_text(item.get("value_type")) == "number" else "categorical_jaccard_gap",
                )
            )
        weighted_scores: list[tuple[float, float]] = []
        completeness_values: list[float] = []
        for code, label, target_value, weight, method in fields:
            competitor_values: list[Any] = []
            for competitor_id in competitor_ids:
                product = context["products"].get(competitor_id, {})
                if code == "price_cny":
                    competitor_values.append(product.get("price_cny"))
                else:
                    param = next((item for item in context["params"].get(competitor_id, []) if clean_text(item.get("field_code")) == code), None)
                    competitor_values.append(param.get("field_value") if param else None)
            present_values = [value for value in competitor_values if value not in (None, "")]
            completeness = (int(target_value not in (None, "")) + len(present_values)) / max(1, 1 + len(competitor_ids))
            gap: float | None = None
            if target_value not in (None, "") and present_values:
                target_number = safe_number(target_value)
                numeric_values = [safe_number(value) for value in present_values]
                if method == "numeric_relative_gap" and target_number is not None and all(value is not None for value in numeric_values):
                    mean_value = sum(value for value in numeric_values if value is not None) / len(numeric_values)
                    gap = min(1.0, abs(target_number - mean_value) / max(abs(target_number), abs(mean_value), 1.0))
                else:
                    target_tokens = set(split_values(target_value)) or {clean_text(target_value)}
                    similarities: list[float] = []
                    for value in present_values:
                        competitor_tokens = set(split_values(value)) or {clean_text(value)}
                        union = target_tokens | competitor_tokens
                        similarities.append(len(target_tokens & competitor_tokens) / len(union) if union else 1.0)
                    gap = 1.0 - (sum(similarities) / len(similarities))
            fit_score = round((1.0 - gap) * completeness * 100, 2) if gap is not None else 0.0
            interpretation = "输入相似度较高" if fit_score >= 75 else "输入存在中等差异" if fit_score >= 40 else "输入差异较大或完整性不足"
            rows.append(
                [case_id, target_id, "|".join(competitor_ids), code, label, target_value,
                 " | ".join("缺失" if value in (None, "") else str(value) for value in competitor_values), method,
                 round(gap, 4) if gap is not None else "", weight, round(completeness * 100, 2), fit_score, interpretation]
            )
            completeness_values.append(completeness)
            weighted_scores.append((fit_score, weight))
        total_weight = sum(weight for _, weight in weighted_scores) or 1.0
        overall = sum(score * weight for score, weight in weighted_scores) / total_weight
        average_completeness = sum(completeness_values) / len(completeness_values) if completeness_values else 0.0
        rows.append(
            [case_id, target_id, "|".join(competitor_ids), "__overall__", "输入仿真拟合度汇总", "", "",
             "weighted_input_similarity", "", total_weight, round(average_completeness * 100, 2), round(overall, 2),
             "仅表示目标产品与所选竞品关键输入的可比程度，不代表真实市场准确率"]
        )
    return rows


def write_validation_result(
    input_path: Path,
    issues: list[ValidationIssue],
    output_path: Path | None = None,
    context: dict[str, Any] | None = None,
) -> Path:
    output = output_path or input_path.with_name(f"{input_path.stem}_校验结果.xlsx")
    workbook = load_workbook(input_path)
    if SHEETS["validation"] in workbook.sheetnames:
        fit_sheet = workbook[SHEETS["validation"]]
        if fit_sheet.max_row > 1:
            fit_sheet.delete_rows(2, fit_sheet.max_row - 1)
    else:
        fit_sheet = workbook.create_sheet(SHEETS["validation"])
        setup_sheet(fit_sheet, VALIDATION_HEADERS, platform=True)
    for row in build_input_fit_rows(context or {}):
        fit_sheet.append(row)
    if SHEETS["input_checks"] in workbook.sheetnames:
        check_sheet = workbook[SHEETS["input_checks"]]
        if check_sheet.max_row > 1:
            check_sheet.delete_rows(2, check_sheet.max_row - 1)
    else:
        check_sheet = workbook.create_sheet(SHEETS["input_checks"])
        setup_sheet(check_sheet, INPUT_CHECK_HEADERS, platform=True)
    for issue in issues:
        check_sheet.append(issue.as_row())
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def convert_param_value(row: dict[str, Any]) -> Any:
    value = row.get("field_value")
    value_type = clean_text(row.get("value_type"))
    if clean_text(row.get("data_status")) == "missing":
        return None
    if value_type == "number":
        number = safe_number(value)
        return int(number) if number is not None and number.is_integer() else number
    if value_type == "boolean":
        return truthy(value)
    return clean_text(value)


def evidence_payload(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {key: row.get(key) for key in EVIDENCE_HEADERS if row.get(key) not in (None, "")}
        for row in rows
    ]


def build_product_payload(record_id: str, context: dict[str, Any], *, custom_competitor: bool = False, competitor_id: int = -1) -> dict[str, Any]:
    row = context["products"][record_id]
    specs: dict[str, Any] = {}
    params: list[dict[str, Any]] = []
    category = clean_text(row.get("category"))
    subcategory = clean_text(row.get("subcategory"))
    for index, param in enumerate(context["params"].get(record_id, []), 1):
        code = clean_text(param.get("field_code"))
        value = convert_param_value(param)
        if value is not None:
            specs[code] = value
        dictionary_row = context["dictionary"].get((category, subcategory, code), {})
        params.append(
            {
                "id": f"param_{index}", "name": code, "raw_name": code,
                "label": clean_text(param.get("field_name_cn") or dictionary_row.get("field_name_cn") or code),
                "value": value, "enabled": value is not None, "locked": False, "is_locked": False,
                "is_preset": bool(dictionary_row), "is_custom": not bool(dictionary_row),
                "field_type": clean_text(param.get("value_type") or dictionary_row.get("field_type") or "string"),
                "unit": clean_text(param.get("unit") or dictionary_row.get("unit")),
                "default_weight": safe_number(dictionary_row.get("default_weight")) or 1.0,
                "weight": safe_number(dictionary_row.get("default_weight")) or 1.0,
            }
        )
    all_evidence: list[dict[str, Any]] = []
    for (product_id, _), rows in context["evidence"].items():
        if product_id == record_id:
            all_evidence.extend(evidence_payload(rows))
    price = safe_number(row.get("price_cny"))
    payload: dict[str, Any] = {
        "product_name": clean_text(row.get("product_name")), "brand": clean_text(row.get("brand")),
        "confirmed_sku": clean_text(row.get("confirmed_sku")), "category": category, "subcategory": subcategory,
        "variant_description": clean_text(row.get("variant_description")), "price_cny": price,
        "price_status": clean_text(row.get("price_status")), "specifications": specs, "params": params,
        "data_provenance": {
            "schema_version": SCHEMA_VERSION, "product_record_id": record_id,
            "collection_date": clean_text(row.get("collection_date")), "collector": clean_text(row.get("collector")),
            "review_status": clean_text(row.get("review_status")),
            "price_type": clean_text(row.get("price_type")),
            "sales_channel": clean_text(row.get("sales_channel")),
            "product_url": clean_text(row.get("product_url")),
            "evidence": all_evidence,
        },
    }
    if custom_competitor:
        payload.update({"id": competitor_id, "is_custom": True, "source": "custom", "competitor_type": "custom"})
    return payload


def build_market_profile(profile_row: dict[str, Any], competitors: list[dict[str, Any]], assumed_count: int) -> dict[str, Any]:
    crowd_profile = json_value(profile_row.get("crowd_profile"), {})
    crowd_segments = json_value(profile_row.get("crowd_segments"), [])
    strategies = json_value(profile_row.get("strategies"), [])
    scenes = json_value(profile_row.get("scenes"), [])
    target_crowd = clean_text(profile_row.get("target_crowd"))
    if not crowd_segments:
        crowd_segments = [
            {
                "name": target_crowd or "目标用户",
                "ratio": 100,
                "is_custom": True,
                "profile": crowd_profile,
            }
        ]
    return {
        "target_crowd": target_crowd, "crowd_profile": crowd_profile,
        "crowd_segments": crowd_segments, "strategy": strategies[0] if strategies else "",
        "strategies": strategies, "strategy_details": json_value(profile_row.get("strategy_details"), {}),
        "scene": scenes[0] if scenes else "", "scenes": scenes,
        "scene_detail": json_value(profile_row.get("scene_details"), {}).get(scenes[0], {}) if scenes else {},
        "scene_details": json_value(profile_row.get("scene_details"), {}),
        "sample_size": int(safe_number(profile_row.get("sample_size")) or 10000),
        "competitors": competitors,
        "market_assumptions": {"assumed_market_competitor_count": assumed_count},
        "decision_weight_profile": json_value(profile_row.get("decision_weight_profile"), {"template": "default"}),
        "social_propagation_config": json_value(profile_row.get("social_propagation_config"), {}),
    }


def compile_cases(context: dict[str, Any], client: "ApiClient | None" = None) -> list[dict[str, Any]]:
    compiled: list[dict[str, Any]] = []
    for case in context["cases"]:
        case_id = clean_text(case.get("case_id"))
        target_id = clean_text(case.get("target_product_id"))
        target = build_product_payload(target_id, context)
        custom_competitors: list[dict[str, Any]] = []
        for index, competitor_id in enumerate(custom_competitor_ids(case, context), 1):
            custom_competitors.append(build_product_payload(competitor_id, context, custom_competitor=True, competitor_id=-index))
        target_count = int(safe_number(case.get("auto_competitor_count")) or 5)
        library_count = max(0, target_count - len(custom_competitors))
        library_competitors: list[dict[str, Any]] = []
        if library_count:
            if client is None:
                raise RuntimeError(f"{case_id}: 需要连接产品库自动选择{library_count}个竞品")
            library_competitors = client.select_library_competitors(
                target, library_count, exclude_products=custom_competitors
            )
        competitors = custom_competitors + library_competitors
        assumed_count = int(safe_number(case.get("assumed_market_competitor_count")) or 20)
        market = build_market_profile(context["profiles"][case_id], competitors, assumed_count)
        payload = {
            "case_id": case_id, "project_name": clean_text(case.get("project_name")),
            "showcase": is_showcase_case(case),
            "product_definition": target, "market_config": market,
        }
        raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
        payload["input_hash"] = hashlib.sha256(raw.encode("utf-8")).hexdigest()
        compiled.append(payload)
    return compiled


def write_compiled(compiled: list[dict[str, Any]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    index: list[dict[str, Any]] = []
    for item in compiled:
        path = output_dir / f"{item['case_id']}.json"
        path.write_text(json.dumps(item, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        index.append({"case_id": item["case_id"], "input_hash": item["input_hash"], "path": str(path)})
    (output_dir / "index.json").write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")


def json_transport_value(value: Any) -> Any:
    """Normalize workbook/API values before handing them to httpx's strict JSON encoder."""

    def encode_unknown(item: Any) -> Any:
        if isinstance(item, date):
            return item.isoformat()
        if isinstance(item, Decimal):
            return float(item)
        if isinstance(item, Path):
            return str(item)
        raise TypeError(f"Object of type {type(item).__name__} is not JSON serializable")

    return json.loads(json.dumps(value, ensure_ascii=False, default=encode_unknown))


class ApiRequestError(RuntimeError):
    def __init__(self, message: str, *, status_code: int | None = None, payload: Any = None):
        super().__init__(message)
        self.status_code = status_code
        self.payload = payload


class ApiClient:
    def __init__(self, *, require_credentials: bool = True):
        self.base_url = os.getenv("AGENTSIM_API_BASE", "").strip().rstrip("/")
        self.username = os.getenv("AGENTSIM_TEST_USERNAME", DEFAULT_TEST_USERNAME).strip()
        self.password = os.getenv("AGENTSIM_TEST_PASSWORD", "")
        if not self.base_url:
            raise RuntimeError("缺少 AGENTSIM_API_BASE")
        parsed = urlparse(self.base_url)
        if parsed.scheme not in {"http", "https"}:
            raise RuntimeError("AGENTSIM_API_BASE 必须是 http 或 https URL")
        allow_insecure_http = truthy(os.getenv("AGENTSIM_ALLOW_INSECURE_HTTP", ""))
        if parsed.scheme == "http" and (parsed.hostname or "") not in {"localhost", "127.0.0.1", "::1"} and not allow_insecure_http:
            raise RuntimeError("非本地HTTP默认禁用；仅测试环境可显式设置 AGENTSIM_ALLOW_INSECURE_HTTP=true")
        if require_credentials and (not self.username or not self.password):
            raise RuntimeError("缺少 AGENTSIM_TEST_USERNAME 或 AGENTSIM_TEST_PASSWORD")
        self.http = httpx.Client(base_url=self.base_url, timeout=45, follow_redirects=True, trust_env=False)
        self.token = ""
        self.user: dict[str, Any] = {}

    def close(self) -> None:
        self.http.close()

    def health(self) -> dict[str, Any]:
        response = self.http.get("/health")
        if response.status_code >= 400:
            raise ApiRequestError("API健康检查失败", status_code=response.status_code)
        data = response.json()
        if not data.get("ok"):
            raise ApiRequestError("API健康检查未通过", status_code=response.status_code, payload=data)
        return data

    def login(self) -> dict[str, Any]:
        if not self.password:
            raise RuntimeError("缺少 AGENTSIM_TEST_PASSWORD")
        response = self.http.post("/api/auth/login", json={"username": self.username, "password": self.password})
        if response.status_code >= 400:
            raise ApiRequestError("测试账号登录失败", status_code=response.status_code)
        data = response.json()
        token = clean_text(data.get("access_token"))
        if not token:
            raise ApiRequestError("登录响应缺少 access_token", status_code=response.status_code)
        self.token = token
        me = self.request("GET", "/api/auth/me", retry_safe=True, allow_relogin=False)
        actual_username = clean_text(me.get("username"))
        if actual_username != self.username:
            self.token = ""
            raise RuntimeError(f"登录账号不匹配：期望 {self.username}，实际 {actual_username}")
        if clean_text(me.get("plan_type")) != "pro":
            self.token = ""
            raise RuntimeError("专用测试账号不是 Pro，禁止创建批量测试项目")
        self.user = me
        return me

    def request(
        self,
        method: str,
        path: str,
        *,
        retry_safe: bool = False,
        allow_relogin: bool = True,
        **kwargs: Any,
    ) -> Any:
        # Non-idempotent requests are never retried for network/5xx failures. The
        # second slot exists only so a confirmed 401 can re-login and resend once.
        attempts = 5 if retry_safe else 2
        delays = [15, 30, 60, 120, 240]
        relogged = False
        last_error: Exception | None = None
        kwargs = dict(kwargs)
        if "json" in kwargs:
            kwargs["json"] = json_transport_value(kwargs["json"])
        for attempt in range(attempts):
            headers = dict(kwargs.pop("headers", {}) or {})
            if self.token:
                headers["Authorization"] = f"Bearer {self.token}"
            try:
                response = self.http.request(method, path, headers=headers, **kwargs)
            except (httpx.TimeoutException, httpx.NetworkError) as exc:
                last_error = exc
                if not retry_safe or attempt + 1 >= attempts:
                    raise ApiRequestError(f"网络请求失败：{method} {path}") from exc
                time.sleep(delays[attempt])
                continue
            if response.status_code == 401 and allow_relogin and not relogged:
                relogged = True
                self.login()
                continue
            if response.status_code == 429 and retry_safe and attempt + 1 < attempts:
                retry_after = safe_number(response.headers.get("Retry-After"))
                time.sleep(max(1, int(retry_after or delays[attempt])))
                continue
            if response.status_code in {502, 503, 504} and retry_safe and attempt + 1 < attempts:
                time.sleep(delays[attempt])
                continue
            data: Any
            try:
                data = response.json() if response.content else {}
            except ValueError:
                data = {"detail": response.text[:1000]}
            if response.status_code >= 400:
                raise ApiRequestError(
                    f"API请求失败：{method} {path} HTTP {response.status_code}",
                    status_code=response.status_code,
                    payload=data,
                )
            return data
        raise ApiRequestError(f"API请求失败：{method} {path}") from last_error

    def find_project_by_name(self, project_name: str) -> list[dict[str, Any]]:
        matches: list[dict[str, Any]] = []
        for page in range(1, 6):
            data = self.request("GET", "/api/simulations", retry_safe=True, params={"page": page, "page_size": 100})
            items = data.get("items") if isinstance(data.get("items"), list) else []
            matches.extend(item for item in items if clean_text(item.get("project_name")) == project_name)
            if len(items) < 100:
                break
        return matches

    def select_library_competitors(
        self,
        target: dict[str, Any],
        count: int,
        *,
        exclude_products: Iterable[dict[str, Any]] = (),
    ) -> list[dict[str, Any]]:
        data = self.request(
            "GET", "/api/products", retry_safe=True,
            params={"category": target.get("category"), "subcategory": target.get("subcategory"), "limit": 100},
        )
        target_price = safe_number(target.get("price_cny")) or 1.0
        target_name = clean_text(target.get("product_name")).lower()
        target_sku = clean_text(target.get("confirmed_sku")).lower()
        excluded_names = {clean_text(item.get("product_name")).lower() for item in exclude_products}
        excluded_skus = {
            clean_text(item.get("confirmed_sku")).lower()
            for item in exclude_products
            if clean_text(item.get("confirmed_sku"))
        }
        candidates: list[tuple[float, int, dict[str, Any]]] = []
        for item in data.get("items") or []:
            price = safe_number(item.get("price_cny"))
            if price is None or price <= 0:
                continue
            if clean_text(item.get("product_name")).lower() == target_name:
                continue
            if target_sku and clean_text(item.get("confirmed_sku")).lower() == target_sku:
                continue
            if is_same_library_product(target, item):
                continue
            if clean_text(item.get("product_name")).lower() in excluded_names:
                continue
            if clean_text(item.get("confirmed_sku")).lower() in excluded_skus:
                continue
            candidates.append((abs(math.log(price / target_price)), int(item.get("id") or 0), item))
        candidates.sort(key=lambda row: (row[0], row[1]))
        return [item for _, _, item in candidates[: max(1, min(count, 50))]]


def append_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {**state, "updated_at": time.strftime("%Y-%m-%dT%H:%M:%S%z")}
    with path.open("a", encoding="utf-8") as file:
        file.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")


def load_latest_states(path: Path) -> dict[str, dict[str, Any]]:
    states: dict[str, dict[str, Any]] = {}
    if not path.exists():
        return states
    for line in path.read_text(encoding="utf-8").splitlines():
        try:
            item = json.loads(line)
        except json.JSONDecodeError:
            continue
        case_id = clean_text(item.get("case_id"))
        if case_id:
            states[case_id] = item
    return states


def nested(data: Any, *keys: str, default: Any = None) -> Any:
    current = data
    for key in keys:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
    return default if current is None else current


def write_summary(run_dir: Path, states: dict[str, dict[str, Any]], compiled: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "批次汇总"
    headers = [
        "case_id", "test_username", "project_id", "task_id", "产品名称", "品牌", "SKU", "品类", "价格",
        "竞品数量", "状态", "耗时秒", "购买意愿", "仿真环境份额", "全市场情景份额", "RCI",
        "价格覆盖率", "数据缺口", "首选策略", "策略仿真ROI", "差异化审计", "质量警告",
        "错误信息", "报告路径",
    ]
    setup_sheet(sheet, headers)
    for case in compiled:
        state = states.get(case["case_id"], {})
        report_path = Path(clean_text(state.get("report_path"))) if state.get("report_path") else None
        report: dict[str, Any] = {}
        if report_path and report_path.exists():
            report = json.loads(report_path.read_text(encoding="utf-8"))
        product = case["product_definition"]
        market = case["market_config"]
        strategies = nested(report, "chart_data", "strategy_roi", default=[]) or []
        best_strategy = max(strategies, key=lambda item: float(item.get("roi") or 0), default={}) if isinstance(strategies, list) else {}
        audit = report.get("differentiation_audit") or nested(report, "chart_data", "differentiation_audit", default={})
        gaps = report.get("data_gaps") or nested(report, "chart_data", "data_gaps", default={})
        scope = report.get("market_share_scope") or nested(report, "chart_data", "market_share_scope", default={})
        sheet.append(
            [
                case["case_id"], state.get("test_username"), state.get("project_id"), state.get("task_id"),
                product.get("product_name"), product.get("brand"), product.get("confirmed_sku"),
                f"{product.get('category')}/{product.get('subcategory')}", product.get("price_cny"),
                len(market.get("competitors") or []), state.get("status"), state.get("duration_seconds"),
                nested(report, "aggregation", "purchase_intent_avg"), scope.get("simulation_environment_share"),
                scope.get("full_market_scenario_share"), scope.get("relative_competitiveness_index"),
                gaps.get("price_coverage_pct"), json.dumps(gaps.get("missing_items") or [], ensure_ascii=False),
                best_strategy.get("name"), best_strategy.get("roi"), json.dumps(audit or {}, ensure_ascii=False),
                json.dumps(report.get("quality_warnings") or [], ensure_ascii=False), state.get("error_message"),
                str(report_path) if report_path else "",
            ]
        )
    workbook.save(run_dir / "summary.xlsx")


def project_name(batch_id: str, case: dict[str, Any]) -> str:
    prefix = f"[TEST-BATCH:{batch_id}:{case['case_id']}]"
    showcase_prefix = "【代表案例】" if bool(case.get("showcase")) else ""
    return f"{showcase_prefix}{prefix} {case['project_name']}"[:160]


def has_expected_batch_prefix(project_name_value: Any, batch_id: str, case_id: str) -> bool:
    """Accept normal and representative-case names while retaining strict batch ownership."""
    expected_prefix = f"[TEST-BATCH:{batch_id}:{case_id}]"
    name = clean_text(project_name_value)
    return name.startswith(expected_prefix) or name.startswith(f"【代表案例】{expected_prefix}")


def state_from_project(state: dict[str, Any], project: dict[str, Any]) -> dict[str, Any]:
    return {
        **state,
        "project_id": project.get("id") or state.get("project_id"),
        "task_id": project.get("task_id") or state.get("task_id"),
        "status": clean_text(project.get("status")) or state.get("status"),
        "error_code": project.get("error_code"),
        "error_message": project.get("error_reason"),
    }


def run_one_case(
    client: ApiClient,
    case: dict[str, Any],
    *,
    batch_id: str,
    run_dir: Path,
    state_path: Path,
    previous: dict[str, Any] | None,
    poll_interval: int,
    timeout_seconds: int,
) -> dict[str, Any]:
    state = {
        "case_id": case["case_id"], "test_username": client.username, "test_user_id": client.user.get("id"),
        "ownership_verified": False, "input_hash": case["input_hash"],
        "status": "validated", "project_id": None, "task_id": None, "error_code": None,
        "error_message": None, "report_path": None,
    }
    if previous:
        if clean_text(previous.get("test_username")) != client.username:
            raise RuntimeError(f"{case['case_id']}: 状态文件属于其他账号，禁止跨账号恢复")
        if clean_text(previous.get("input_hash")) != case["input_hash"]:
            raise RuntimeError(f"{case['case_id']}: 输入已改变，禁止复用旧任务")
        state.update(previous)
        if state.get("status") == "completed" and state.get("report_path") and Path(str(state["report_path"])).exists():
            return state

    started = time.monotonic()
    name = project_name(batch_id, case)
    project: dict[str, Any] = {}
    try:
        if state.get("project_id"):
            project = client.request("GET", f"/api/simulations/{int(state['project_id'])}", retry_safe=True)
            state = state_from_project(state, project)
        else:
            try:
                project = client.request("POST", "/api/simulations", json={"project_name": name})
            except ApiRequestError as exc:
                matches = client.find_project_by_name(name)
                if len(matches) != 1:
                    state.update({"status": "needs_review", "error_code": "CREATE_UNCERTAIN", "error_message": str(exc)})
                    append_state(state_path, state)
                    return state
                project = client.request("GET", f"/api/simulations/{int(matches[0]['id'])}", retry_safe=True)
            state = state_from_project({**state, "status": "created"}, project)
            append_state(state_path, state)

        if int(project.get("user_id") or 0) != int(client.user.get("id") or -1):
            raise RuntimeError(f"{case['case_id']}: 项目账号归属校验失败，已停止")
        state["ownership_verified"] = True
        state["test_user_id"] = client.user.get("id")
        append_state(state_path, state)

        if clean_text(project.get("status")) == "completed":
            report_response = client.request("GET", f"/api/simulations/{int(project['id'])}/report", retry_safe=True)
            report = report_response.get("report") or report_response.get("result_data") or {}
            report_path = run_dir / "reports" / f"{case['case_id']}.json"
            report_path.parent.mkdir(parents=True, exist_ok=True)
            report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            state.update({"status": "completed", "report_path": str(report_path)})
            append_state(state_path, state)
            return state

        if not project.get("task_id"):
            client.request("PUT", f"/api/simulations/{int(project['id'])}/step1", retry_safe=True, json={"product_definition": case["product_definition"]})
            client.request("PUT", f"/api/simulations/{int(project['id'])}/step2", retry_safe=True, json={"market_config": case["market_config"]})
            state["status"] = "configured"
            append_state(state_path, state)
            try:
                project = client.request("POST", f"/api/simulations/{int(project['id'])}/submit", json={})
            except ApiRequestError:
                project = client.request("GET", f"/api/simulations/{int(project['id'])}", retry_safe=True)
                if clean_text(project.get("status")) not in {"submitted", "running", "report_generation_waiting", "completed"}:
                    raise
            state = state_from_project({**state, "status": "submitted"}, project)
            append_state(state_path, state)
            if not project.get("task_id"):
                try:
                    run_response = client.request("POST", f"/api/simulations/{int(project['id'])}/run", json={})
                    project = run_response.get("project") or project
                    task = run_response.get("task") or {}
                    state["task_id"] = task.get("task_id") or project.get("task_id")
                except ApiRequestError:
                    project = client.request("GET", f"/api/simulations/{int(project['id'])}", retry_safe=True)
                    if not project.get("task_id"):
                        state.update({"status": "needs_review", "error_code": "RUN_UNCERTAIN", "error_message": "启动请求结果不确定"})
                        append_state(state_path, state)
                        return state
                state = state_from_project({**state, "status": "queued"}, project)
                append_state(state_path, state)

        deadline = time.monotonic() + timeout_seconds
        consecutive_failures = 0
        while time.monotonic() < deadline:
            try:
                progress = client.request("GET", f"/api/simulations/{int(state['project_id'])}/progress", retry_safe=True)
                consecutive_failures = 0
            except ApiRequestError:
                consecutive_failures += 1
                if consecutive_failures >= 5:
                    raise RuntimeError("连续5次无法读取任务进度，批次暂停")
                time.sleep(poll_interval)
                continue
            project = progress.get("project") or {}
            task = progress.get("task") or {}
            status = clean_text(project.get("status") or task.get("status"))
            state.update(
                {
                    "status": status or state.get("status"), "task_id": task.get("task_id") or project.get("task_id") or state.get("task_id"),
                    "last_stage": task.get("stage"), "percent": task.get("percent"),
                    "error_code": project.get("error_code"), "error_message": project.get("error_reason"),
                }
            )
            append_state(state_path, state)
            if status == "completed":
                report_response = client.request("GET", f"/api/simulations/{int(state['project_id'])}/report", retry_safe=True)
                report = report_response.get("report") or report_response.get("result_data") or {}
                report_path = run_dir / "reports" / f"{case['case_id']}.json"
                report_path.parent.mkdir(parents=True, exist_ok=True)
                report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
                state.update({"status": "completed", "report_path": str(report_path), "duration_seconds": round(time.monotonic() - started, 1)})
                append_state(state_path, state)
                return state
            if status in {"failed", "cancelled"}:
                state["duration_seconds"] = round(time.monotonic() - started, 1)
                append_state(state_path, state)
                return state
            time.sleep(poll_interval)
        state.update({"status": "needs_review", "error_code": "LOCAL_WAIT_TIMEOUT", "error_message": "本地等待超过限制，可使用resume继续"})
        append_state(state_path, state)
        return state
    except Exception as exc:
        state.update({"status": "failed", "error_code": state.get("error_code") or "RUNNER_ERROR", "error_message": str(exc), "duration_seconds": round(time.monotonic() - started, 1)})
        append_state(state_path, state)
        return state


def prepare_run(input_path: Path, run_dir: Path, batch_id: str, client: ApiClient) -> tuple[list[dict[str, Any]], Path]:
    issues, context = validate_workbook(input_path)
    validation_path = write_validation_result(input_path, issues, run_dir / "validation_result.xlsx", context)
    errors = [issue for issue in issues if issue.severity == "error"]
    if errors:
        raise RuntimeError(f"Excel校验失败：{len(errors)}个错误，详见 {validation_path}")
    compiled = compile_cases(context, client)
    write_compiled(compiled, run_dir / "compiled_payloads")
    copied_input = run_dir / "input_original.xlsx"
    if input_path.resolve() != copied_input.resolve():
        shutil.copy2(input_path, copied_input)
    meta = {
        "batch_id": batch_id, "input_path": str(copied_input), "test_username": client.username,
        "test_user_id": client.user.get("id"), "schema_version": SCHEMA_VERSION,
        "server_results_persisted": True, "cleanup_requires_explicit_confirmation": True,
    }
    (run_dir / "batch_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    return compiled, run_dir / "run_state.jsonl"


def execute_batch(
    input_path: Path,
    *,
    batch_id: str,
    run_dir: Path,
    max_items: int,
    poll_interval: int,
    cooldown: int,
    timeout_seconds: int,
    case_ids: list[str] | None = None,
) -> int:
    client = ApiClient()
    try:
        client.health()
        client.login()
        run_dir.mkdir(parents=True, exist_ok=True)
        compiled, state_path = prepare_run(input_path, run_dir, batch_id, client)
        if case_ids:
            requested = [clean_text(case_id) for case_id in case_ids if clean_text(case_id)]
            available = {case["case_id"] for case in compiled}
            missing = [case_id for case_id in requested if case_id not in available]
            if missing:
                raise RuntimeError(f"指定case_id不存在：{'|'.join(missing)}")
            requested_set = set(requested)
            compiled = [case for case in compiled if case["case_id"] in requested_set]
        if max_items > 0:
            compiled = compiled[:max_items]
        states = load_latest_states(state_path)
        for index, case in enumerate(compiled):
            result = run_one_case(
                client, case, batch_id=batch_id, run_dir=run_dir, state_path=state_path,
                previous=states.get(case["case_id"]), poll_interval=poll_interval, timeout_seconds=timeout_seconds,
            )
            states[case["case_id"]] = result
            write_summary(run_dir, states, compiled)
            if result.get("error_message") == "连续5次无法读取任务进度，批次暂停":
                break
            if index + 1 < len(compiled) and cooldown > 0:
                time.sleep(cooldown)
        write_summary(run_dir, states, compiled)
        completed = sum(1 for case in compiled if states.get(case["case_id"], {}).get("status") == "completed")
        print(json.dumps({"batch_id": batch_id, "run_dir": str(run_dir), "total": len(compiled), "completed": completed}, ensure_ascii=False))
        return 0 if completed == len(compiled) else 1
    finally:
        client.close()


def command_template(args: argparse.Namespace) -> int:
    output = Path(args.output)
    write_workbook(create_workbook(), output)
    print(json.dumps({"output": str(output), "schema_version": SCHEMA_VERSION}, ensure_ascii=False))
    return 0


def command_sample(args: argparse.Namespace) -> int:
    output = Path(args.output)
    write_workbook(create_workbook(include_samples=True, sample_count=args.count, seed=args.seed), output)
    print(json.dumps({"output": str(output), "items": args.count, "seed": args.seed, "synthetic": True}, ensure_ascii=False))
    return 0


def command_validate(args: argparse.Namespace) -> int:
    input_path = Path(args.input)
    issues, context = validate_workbook(input_path)
    output = write_validation_result(input_path, issues, Path(args.output) if args.output else None, context)
    counts = {level: sum(1 for issue in issues if issue.severity == level) for level in ("error", "warning", "info")}
    print(json.dumps({"input": str(input_path), "output": str(output), "enabled_cases": len(context.get("cases", [])), **counts}, ensure_ascii=False))
    return 1 if counts["error"] else 0


def command_compile(args: argparse.Namespace) -> int:
    input_path = Path(args.input)
    issues, context = validate_workbook(input_path)
    errors = [issue for issue in issues if issue.severity == "error"]
    validation_output = write_validation_result(input_path, issues, context=context)
    if errors:
        print(json.dumps({"error": "validation_failed", "count": len(errors), "validation_output": str(validation_output)}, ensure_ascii=False))
        return 1
    needs_library = any(
        len(custom_competitor_ids(case, context))
        < int(safe_number(case.get("auto_competitor_count")) or 5)
        for case in context["cases"]
    )
    client: ApiClient | None = None
    try:
        if needs_library:
            client = ApiClient(require_credentials=False)
            client.health()
        compiled = compile_cases(context, client)
        output_dir = Path(args.output)
        write_compiled(compiled, output_dir)
        print(json.dumps({"output": str(output_dir), "items": len(compiled), "validation_output": str(validation_output)}, ensure_ascii=False))
        return 0
    finally:
        if client:
            client.close()


def command_login_check(_: argparse.Namespace) -> int:
    client = ApiClient()
    try:
        health = client.health()
        user = client.login()
        payload = {
            "api_ok": bool(health.get("ok")), "username": user.get("username"), "user_id": user.get("id"),
            "plan_type": user.get("plan_type"), "login_verified": True,
        }
        print(json.dumps(payload, ensure_ascii=False))
        return 0
    finally:
        client.close()


def command_run(args: argparse.Namespace) -> int:
    batch_id = clean_text(args.batch_id)
    if not re.fullmatch(r"[A-Za-z0-9_-]+", batch_id):
        raise RuntimeError("batch-id只能包含字母、数字、下划线和短横线")
    run_dir = Path(args.run_dir) if args.run_dir else PROJECT_ROOT / "batch_runs" / batch_id
    return execute_batch(
        Path(args.input), batch_id=batch_id, run_dir=run_dir, max_items=args.max_items,
        poll_interval=args.poll_interval, cooldown=args.cooldown, timeout_seconds=args.timeout_seconds,
        case_ids=args.case_ids,
    )


def command_resume(args: argparse.Namespace) -> int:
    run_dir = Path(args.run_dir)
    meta_path = run_dir / "batch_meta.json"
    if not meta_path.exists():
        raise RuntimeError("run-dir缺少batch_meta.json，无法恢复")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    expected_username = os.getenv("AGENTSIM_TEST_USERNAME", DEFAULT_TEST_USERNAME).strip()
    if clean_text(meta.get("test_username")) != expected_username:
        raise RuntimeError("批次属于其他测试账号，禁止跨账号恢复")
    input_path = Path(clean_text(meta.get("input_path")))
    return execute_batch(
        input_path, batch_id=clean_text(meta.get("batch_id")), run_dir=run_dir, max_items=args.max_items,
        poll_interval=args.poll_interval, cooldown=args.cooldown, timeout_seconds=args.timeout_seconds,
        case_ids=args.case_ids,
    )


def command_cleanup(args: argparse.Namespace) -> int:
    run_dir = Path(args.run_dir)
    meta_path = run_dir / "batch_meta.json"
    state_path = run_dir / "run_state.jsonl"
    if not meta_path.exists() or not state_path.exists():
        raise RuntimeError("run-dir缺少batch_meta.json或run_state.jsonl")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    batch_id = clean_text(meta.get("batch_id"))
    if clean_text(args.confirm_batch_id) != batch_id:
        raise RuntimeError("确认批次编号不匹配，未删除任何项目")
    client = ApiClient()
    results: list[dict[str, Any]] = []
    try:
        client.health()
        client.login()
        if clean_text(meta.get("test_username")) != client.username or int(meta.get("test_user_id") or 0) != int(client.user.get("id") or -1):
            raise RuntimeError("当前登录账号与批次归属不一致，禁止清理")
        for case_id, state in load_latest_states(state_path).items():
            project_id = int(state.get("project_id") or 0)
            if not project_id:
                results.append({"case_id": case_id, "status": "skipped", "reason": "no_project_id"})
                continue
            try:
                project = client.request("GET", f"/api/simulations/{project_id}", retry_safe=True)
            except ApiRequestError as exc:
                if exc.status_code == 404:
                    results.append({"case_id": case_id, "project_id": project_id, "status": "already_deleted"})
                    continue
                raise
            if int(project.get("user_id") or 0) != int(client.user.get("id") or -1):
                raise RuntimeError(f"{case_id}: 项目不属于当前测试账号，停止清理")
            if not has_expected_batch_prefix(project.get("project_name"), batch_id, case_id):
                raise RuntimeError(f"{case_id}: 项目名称缺少测试批次前缀，停止清理")
            if clean_text(project.get("status")) in {"running", "submitted", "queued", "report_generation_waiting"}:
                results.append({"case_id": case_id, "project_id": project_id, "status": "skipped", "reason": "active_project"})
                continue
            client.request("DELETE", f"/api/simulations/{project_id}")
            results.append({"case_id": case_id, "project_id": project_id, "status": "deleted"})
        output = run_dir / "cleanup_results.json"
        output.write_text(json.dumps({"batch_id": batch_id, "test_username": client.username, "items": results}, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps({"batch_id": batch_id, "deleted": sum(item["status"] == "deleted" for item in results), "output": str(output)}, ensure_ascii=False))
        return 0
    finally:
        client.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="AgentSim产品Excel与专用测试账号串行仿真工具")
    subparsers = parser.add_subparsers(dest="command", required=True)

    template_parser = subparsers.add_parser("template", help="生成空白产品信息采集模板")
    template_parser.add_argument("--output", default="AgentSim批量产品信息采集模板_v3.xlsx")
    template_parser.set_defaults(handler=command_template)

    sample_parser = subparsers.add_parser("sample", help="生成1～2组全字段模拟数据")
    sample_parser.add_argument("--output", default="AgentSim模拟产品数据_v3.xlsx")
    sample_parser.add_argument("--count", type=int, choices=(1, 2), default=2)
    sample_parser.add_argument("--seed", type=int, default=20260804)
    sample_parser.set_defaults(handler=command_sample)

    validate_parser = subparsers.add_parser("validate", help="离线校验Excel")
    validate_parser.add_argument("--input", required=True)
    validate_parser.add_argument("--output", default="")
    validate_parser.set_defaults(handler=command_validate)

    compile_parser = subparsers.add_parser("compile", help="将Excel转换为API JSON")
    compile_parser.add_argument("--input", required=True)
    compile_parser.add_argument("--output", required=True)
    compile_parser.set_defaults(handler=command_compile)

    login_parser = subparsers.add_parser("login-check", help="验证API和专用Pro测试账号")
    login_parser.set_defaults(handler=command_login_check)

    def add_run_arguments(command_parser: argparse.ArgumentParser, *, include_input: bool) -> None:
        if include_input:
            command_parser.add_argument("--input", required=True)
            command_parser.add_argument("--batch-id", required=True)
            command_parser.add_argument("--run-dir", default="")
        else:
            command_parser.add_argument("--run-dir", required=True)
        command_parser.add_argument("--max-items", type=int, default=0, help="0表示全部")
        command_parser.add_argument(
            "--case-id",
            dest="case_ids",
            action="append",
            default=[],
            help="仅运行指定case_id，可重复传入；为空时按工作簿全部任务处理",
        )
        command_parser.add_argument("--poll-interval", type=int, default=15)
        command_parser.add_argument("--cooldown", type=int, default=20)
        command_parser.add_argument("--timeout-seconds", type=int, default=10800)

    run_parser = subparsers.add_parser("run", help="使用专用账号串行运行")
    add_run_arguments(run_parser, include_input=True)
    run_parser.set_defaults(handler=command_run)

    resume_parser = subparsers.add_parser("resume", help="恢复中断批次")
    add_run_arguments(resume_parser, include_input=False)
    resume_parser.set_defaults(handler=command_resume)

    cleanup_parser = subparsers.add_parser("cleanup", help="删除专用账号下已结束的指定测试批次项目")
    cleanup_parser.add_argument("--run-dir", required=True)
    cleanup_parser.add_argument("--confirm-batch-id", required=True, help="必须与批次元数据完全一致")
    cleanup_parser.set_defaults(handler=command_cleanup)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        return int(args.handler(args))
    except (RuntimeError, ApiRequestError, json.JSONDecodeError, OSError, ValueError) as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
