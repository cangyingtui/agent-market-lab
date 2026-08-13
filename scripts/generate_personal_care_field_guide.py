from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


CASE_MAP = {
    "WATERPIK-WP660": ("pc_waterpik_wp660", "关注口腔清洁、牙龈护理和家庭使用便利性的消费者", ["测评背书", "卖点强化", "内容种草"], ["家庭口腔护理", "综合电商"]),
    "HYPERICE-HV2PRO": ("pc_hyperice_hv2pro", "关注运动恢复、肌肉放松和专业性能的健身人群", ["专业测评", "场景体验", "卖点强化"], ["运动恢复", "综合电商"]),
    "TRIPOLLAR-STOPVX2": ("pc_tripollar_stopvx2", "关注居家抗衰、射频护理和安全性的中高消费人群", ["专业科普", "测评背书", "场景包装"], ["居家美容护理", "内容种草"]),
    "ULIKE-AIR10": ("pc_ulike_air10", "关注家用脱毛效率、舒适度和隐私性的年轻消费者", ["内容种草", "效果测评", "场景包装"], ["居家脱毛", "短视频电商"]),
    "YUWELL-YHW2": ("pc_yuwell_yhw2", "关注家庭体温监测、老人儿童照护和操作便捷性的家庭用户", ["专业科普", "使用演示", "渠道信任"], ["家庭健康监测", "综合电商"]),
    "YUWELL-YE680A": ("pc_yuwell_ye680a", "关注居家血压管理、操作便捷性和品牌可信度的中老年家庭", ["专业科普", "使用演示", "渠道信任"], ["慢病家庭管理", "综合电商"]),
    "SINOCARE-GA3": ("pc_sinocare_ga3", "关注居家血糖监测、耗材成本和检测便捷性的慢病管理人群", ["专业科普", "使用演示", "性价比沟通"], ["居家血糖管理", "综合电商"]),
}


def sheet_rows(workbook, index: int) -> list[dict]:
    worksheet = workbook.worksheets[index]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    return [dict(zip(headers, row)) for row in iterator if any(value not in (None, "") for value in row)]


def setup(worksheet, headers: list[str], widths: list[int]) -> None:
    worksheet.append(headers)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9E2F3")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.sheet_view.showGridLines = False
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_rows(worksheet, values: list[tuple], fill: PatternFill | None = None) -> None:
    for value in values:
        worksheet.append(value)
        if fill:
            for cell in worksheet[worksheet.max_row]:
                cell.fill = fill


def main() -> int:
    if len(sys.argv) != 3:
        raise SystemExit("usage: generate_personal_care_field_guide.py INPUT OUTPUT")
    source = Path(sys.argv[1])
    output = Path(sys.argv[2])
    source_book = load_workbook(source, data_only=True, read_only=True)
    products = sheet_rows(source_book, 2)
    params = sheet_rows(source_book, 3)
    evidence = sheet_rows(source_book, 4)
    dictionary = sheet_rows(source_book, 5)
    dictionary_map = {
        (str(row.get("category") or "").strip(), str(row.get("subcategory") or "").strip(), str(row.get("field_code") or "").strip()): row
        for row in dictionary
    }
    evidence_fields = {
        (str(row.get("product_record_id") or "").strip(), str(row.get("field_code") or "").strip())
        for row in evidence
    }

    workbook = Workbook()
    workbook.remove(workbook.active)
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    suggested_fill = PatternFill("solid", fgColor="DDEBF7")
    optional_fill = PatternFill("solid", fgColor="E2F0D9")
    warning_fill = PatternFill("solid", fgColor="FCE4D6")

    worksheet = workbook.create_sheet("01_字段必填说明")
    setup(worksheet, ["工作表", "字段", "级别", "何时需要", "如何填写", "示例/默认", "留空后果"], [18, 34, 18, 24, 52, 48, 42])
    required = [
        ("01_测试任务", "case_id", "必填（本地标识）", "每个启用任务", "自行填写唯一英文/数字编号；不是服务器project_id", "pc_waterpik_wp660", "无法关联06配置、断点和结果文件"),
        ("01_测试任务", "enabled", "必填", "要运行的任务", "填 是/true/1", "是", "空值表示不运行，服务器不会创建项目"),
        ("01_测试任务", "project_name", "必填", "每个启用任务", "填写用户可见项目名", "Waterpik WP-660市场测试", "服务器创建项目时使用"),
        ("01_测试任务", "target_product_id", "必填", "每个启用任务", "引用02表product_record_id", "WATERPIK-WP660", "找不到目标产品"),
        ("01_测试任务", "competitor_mode", "无需填写", "旧模板兼容字段", "保持为空即可", "", "脚本根据是否填写竞品ID自动判断"),
        ("01_测试任务", "competitor_product_ids", "可选", "仅有自定义竞品时", "用|分隔02表中的竞品ID；留空时全部自动检索", "COMP-A|COMP-B", "填写后保留自定义竞品，并由产品库补足目标数量"),
        ("01_测试任务", "assumed_market_competitor_count", "可选", "需要调整份额情景时", "填写5～50的全市场情景竞品数；留空默认20", "20", "自动使用20"),
        ("06_运行配置", "case_id", "必填", "每个启用任务", "必须与01表完全一致", "pc_waterpik_wp660", "无法关联任务"),
        ("06_运行配置", "target_crowd", "必填", "每个启用任务", "一句话描述目标消费者", "关注家庭口腔护理的消费者", "校验失败"),
        ("06_运行配置", "strategies", "必填", "每个启用任务", "JSON数组；策略可自由设置", "[\"测评背书\",\"卖点强化\"]", "校验失败"),
        ("06_运行配置", "scenes", "必填", "每个启用任务", "JSON数组", "[\"家庭护理\",\"综合电商\"]", "校验失败"),
        ("02_产品主表", "产品核心字段", "必填", "进入任务的产品", "保留现有product_record_id/name/category/subcategory等", "WATERPIK-WP660", "无法组装产品"),
        ("02_产品主表", "price_cny/price_type/price_status", "目标产品必填", "作为待测本品", "价格大于0并标记confirmed/estimated/manual", "674.93 / estimated", "目标产品无法提交"),
        ("04_证据来源", "price_cny证据", "目标产品必填", "每个启用目标产品", "至少一条field_code=price_cny证据", "TriPollar当前需要补", "任务校验失败"),
    ]
    append_rows(worksheet, required, required_fill)
    suggested = [
        ("01_测试任务", "auto_competitor_count", "可选", "每个启用任务", "竞品总目标数量；留空默认5", "5", "自定义竞品不足该数量时由产品库补足"),
        ("06_运行配置", "crowd_profile", "强烈建议", "需要有解释力的结果", "JSON对象，建议填写价格敏感度、功能和渠道偏好", "{\"price_sensitivity\":\"medium\"}", "可留{}，但解释力下降"),
        ("06_运行配置", "strategy_details", "建议", "需要渠道差异和成本分析", "为策略配置actions/channels/economics", "{}", "可运行，但渠道审计可能显示输入不足"),
        ("02_产品主表", "review_status", "规范建议", "所有产品", "建议使用pending/reviewed/rejected", "pending", "当前代码仅检查非空，但现值不符合模板枚举"),
        ("04_证据来源", "参数字段级证据", "可选溯源", "需要逐参数展示来源时", "证据field_code与参数field_code相同", "attachments", "不阻止运行，也不影响当前输入拟合度；只影响逐字段来源展示"),
    ]
    append_rows(worksheet, suggested, suggested_fill)
    optional = [
        ("06_运行配置", "crowd_segments", "可选", "需要多人群分层", "JSON数组", "[]", "默认单一整体人群"),
        ("06_运行配置", "scene_details", "可选", "需要传播裂变标签", "JSON对象", "{}", "使用基础场景"),
        ("06_运行配置", "sample_size", "可选", "每个任务", "正整数；默认10000", "10000", "自动使用10000"),
        ("06_运行配置", "decision_weight_profile", "可选", "需要渠道模板", "JSON对象", "{\"template\":\"default\"}", "自动使用default"),
        ("06_运行配置", "social_propagation_config", "可选", "需要自定义传播参数", "JSON对象", "{}", "使用系统默认"),
        ("03_参数明细", "field_value", "条件可空", "data_status=missing", "确实未知时留空并标记missing", "stall_force", "该参数不参与评分"),
    ]
    append_rows(worksheet, optional, optional_fill)

    worksheet = workbook.create_sheet("02_任务建议")
    setup(worksheet, ["case_id", "enabled", "project_name", "target_product_id", "competitor_mode", "competitor_product_ids", "auto_competitor_count", "assumed_market_competitor_count", "说明"], [28, 10, 40, 24, 20, 28, 22, 30, 46])
    for product in products:
        product_id = str(product.get("product_record_id"))
        worksheet.append([CASE_MAP[product_id][0], "是", product.get("product_name"), product_id, "", "", 5, 20, "无自定义竞品；留空后自动检索同小类产品库竞品"])

    worksheet = workbook.create_sheet("03_运行配置建议")
    setup(worksheet, ["case_id", "target_crowd", "crowd_profile", "crowd_segments", "strategies", "strategy_details", "scenes", "scene_details", "sample_size", "decision_weight_profile", "social_propagation_config"], [28, 46, 70, 18, 42, 75, 38, 60, 14, 30, 48])
    for product in products:
        product_id = str(product.get("product_record_id"))
        case_id, crowd, strategies, scenes = CASE_MAP[product_id]
        profile = {"name": crowd, "price_sensitivity": "medium", "feature_priorities": ["核心效果", "使用便利性", "安全/可信度"], "channel_preferences": ["综合电商", "内容平台"], "risk_concerns": ["效果真实性", "长期使用成本"]}
        details = {strategy: {"actions": [f"围绕{strategy}执行轻量测试"], "channels": ["综合电商"]} for strategy in strategies}
        scene_details = {scene: {"place": scene, "scene_tags": [scene]} for scene in scenes}
        worksheet.append([case_id, crowd, json.dumps(profile, ensure_ascii=False), "[]", json.dumps(strategies, ensure_ascii=False), json.dumps(details, ensure_ascii=False), json.dumps(scenes, ensure_ascii=False), json.dumps(scene_details, ensure_ascii=False), 10000, '{"template":"default"}', '{"external_traffic_per_round":100,"scene_fission_factor":1.0}'])

    worksheet = workbook.create_sheet("04_逐产品缺口")
    setup(worksheet, ["产品ID", "产品", "当前硬性缺口", "可保留为空的字段", "建议复核项", "是否可直接运行"], [24, 42, 42, 40, 52, 26])
    for product in products:
        product_id = str(product.get("product_record_id"))
        hard = []
        optional_missing = []
        review = []
        if (product_id, "price_cny") not in evidence_fields:
            hard.append("缺少price_cny证据")
        for param in params:
            if str(param.get("product_record_id") or "").strip() == product_id and param.get("field_value") in (None, ""):
                optional_missing.append(f"{param.get('field_code')}（状态{param.get('data_status')}）")
        if str(product.get("price_status")) == "estimated":
            review.append("价格为estimated，建议复核中国市场口径")
        if str(product.get("review_status")) not in {"pending", "reviewed", "rejected"}:
            review.append("review_status改为pending或reviewed")
        worksheet.append([product_id, product.get("product_name"), "；".join(hard) or "无（补齐01/06表后）", "；".join(optional_missing) or "无实际缺值", "；".join(review) or "无", "否" if hard else "补齐01/06表后可以"])
        if hard:
            for cell in worksheet[worksheet.max_row]:
                cell.fill = warning_fill

    worksheet = workbook.create_sheet("05_参数逐项判断")
    setup(worksheet, ["产品ID", "小类", "field_code", "参数名称", "当前值", "data_status", "字典是否必填", "默认权重", "可选溯源标注", "处理建议"], [24, 20, 28, 28, 48, 16, 18, 14, 18, 54])
    for product in products:
        product_id = str(product.get("product_record_id"))
        category = str(product.get("category"))
        subcategory = str(product.get("subcategory"))
        for param in params:
            if str(param.get("product_record_id") or "").strip() != product_id:
                continue
            code = str(param.get("field_code") or "").strip()
            dictionary_row = dictionary_map.get((category, subcategory, code), {})
            missing = param.get("field_value") in (None, "")
            if missing and str(param.get("data_status")) == "missing":
                advice = "可保持missing，不参与该参数评分"
            elif (product_id, code) not in evidence_fields:
                advice = "可直接用于拟合；如需逐参数展示出处，可补同field_code证据"
            else:
                advice = "可直接使用"
            worksheet.append([product_id, subcategory, code, param.get("field_name_cn"), param.get("field_value"), param.get("data_status"), dictionary_row.get("required") or "否", dictionary_row.get("default_weight"), "有" if (product_id, code) in evidence_fields else "无", advice])

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
