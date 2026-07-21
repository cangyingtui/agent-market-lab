from __future__ import annotations

import json
from io import BytesIO
from datetime import timedelta
from pathlib import Path
from typing import Any

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database import SessionLocal
from app import export_service
from app.export_service import export_file_path
from app.models import ShareToken, SimulationProject, User
from app.time_utils import utc_now_naive
from engine.distill_client import run_distill_checks_if_enabled
from engine import export_worker


def _sample_report() -> dict[str, Any]:
    return {
        "executive_summary": "pytest 完整报告",
        "target_segments": [{"name": "高端用户", "insight": "关注续航"}],
        "crowd_profile": {
            "name": "高端用户",
            "age_range": "28-45",
            "city_tier": "一线/新一线",
            "income_level": "高收入",
            "price_sensitivity": "low",
            "feature_priorities": ["续航", "屏幕"],
            "channel_preferences": ["品牌旗舰店"],
            "purchase_motivations": ["体验升级"],
            "risk_concerns": ["价格波动"],
        },
        "competitor_insights": [{"source": "product:1", "insight": "竞品价格有优势"}],
        "pricing_analysis": {"summary": "价格可接受", "reference_price": 3999},
        "strategy_recommendations": ["突出防水和续航"],
        "risk_warnings": ["价格数据需要复核"],
        "evidence_used": [{"source": "用户画像数据_pytest", "snippet": "关注性价比"}],
        "agent_samples": [{"agent_id": "agent_001", "segment": "高端用户"}],
        "purchase_decisions": [{"agent_id": "agent_001", "purchase_intent_score": 0.75}],
        "aggregation": {"purchase_intent_avg": 0.75},
        "social_simulation": {
            "rounds_executed": 2,
            "converged": True,
            "node_count": 60,
            "edge_count": 120,
            "average_degree": 4,
        },
        "chart_data": {
            "prompt_version": "chart_data_pytest",
            "plan_type": "pro",
            "overview_metrics": {
                "purchase_intent_index": 75,
                "estimated_market_share": 58.2,
                "target_match": "高度匹配",
                "evidence_count": 3,
                "competitor_count": 1,
            },
            "market_share": [
                {"name": "测试智能手机", "role": "self", "share": 58.2, "value": 58.2},
                {"name": "竞品 A", "role": "competitor", "share": 41.8, "value": 41.8},
            ],
            "purchase_intent_by_segment": [{"name": "高端用户", "value": 75, "count": 1}],
            "param_importance": [{"name": "续航", "importance": 88, "weight": 4}],
            "strategy_roi": [{"name": "突出防水和续航", "roi": 2.1}],
            "channel_effect": [{"name": "电商平台", "share": 100, "value": 100}],
            "price_sensitivity": [{"price": 3599, "intent": 82}, {"price": 3999, "intent": 75}],
            "social_evolution": [
                {"round": 1, "name": "整体人群", "value": 72},
                {"round": 2, "name": "整体人群", "value": 75},
            ],
            "social_rounds": [
                {"round": 1, "overall_purchase_intent": 0.72},
                {"round": 2, "overall_purchase_intent": 0.75},
            ],
            "competitor_radar": {
                "dimensions": ["功能丰富度", "价格竞争力", "品牌影响力", "用户体验", "渠道覆盖"],
                "series": [{"name": "测试智能手机", "role": "self", "values": [80, 70, 60, 75, 65]}],
            },
            "sensitivity_waterfall": [{"name": "当前基线", "delta": 0, "value": 75}],
        },
        "quality_warnings": [],
        "prompt_trace": {"report_builder": {"prompt": "不应导出"}},
        "formal_test_log_path": "logs/formal_runs/private.json",
        "api_key": "sk-should-not-appear",
        "rag_evidence": [{"snippet": "证据", "raw": {"internal": "公开分享不应包含 raw"}}],
    }


def _create_completed_project(client: TestClient, headers: dict[str, str], report: dict[str, Any] | None = None) -> int:
    created = client.post("/api/simulations", headers=headers, json={"project_name": "pytest 导出分享"})
    created.raise_for_status()
    project_id = created.json()["id"]
    with SessionLocal() as db:
        project = db.get(SimulationProject, project_id)
        assert project is not None
        project.status = "completed"
        project.result_data = report or _sample_report()
        project.snapshot_hash = "pytest_snapshot_hash"
        db.commit()
    return project_id


def _promote_to_pro(client: TestClient, headers: dict[str, str]) -> None:
    response = client.get("/api/auth/me", headers=headers)
    response.raise_for_status()
    user_id = response.json()["id"]
    with SessionLocal() as db:
        user = db.get(User, user_id)
        assert user is not None
        user.plan_type = "pro"
        db.commit()


def test_completed_project_can_export_json_and_markdown(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    _promote_to_pro(client, auth_headers)
    project_id = _create_completed_project(client, auth_headers)

    exported = client.post(
        f"/api/simulations/{project_id}/exports",
        headers=auth_headers,
        json={"format": "json"},
    )
    assert exported.status_code == 201
    body = exported.json()
    assert body["status"] == "completed"
    assert body["download_url"] == f"/api/exports/{body['export_task_id']}/download"

    status_response = client.get(f"/api/exports/{body['export_task_id']}", headers=auth_headers)
    assert status_response.status_code == 200
    assert status_response.json()["status"] == "completed"

    downloaded = client.get(body["download_url"], headers=auth_headers)
    assert downloaded.status_code == 200
    payload = json.loads(downloaded.content.decode("utf-8"))
    assert payload["report"]["executive_summary"] == "pytest 完整报告"
    assert payload["report"]["chart_data"]["market_share"][0]["role"] == "self"
    text = downloaded.content.decode("utf-8")
    assert "sk-should-not-appear" not in text
    assert "prompt_trace" not in text
    assert "formal_test_log_path" not in text

    markdown = client.post(
        f"/api/simulations/{project_id}/exports",
        headers=auth_headers,
        json={"format": "markdown"},
    )
    assert markdown.status_code == 201
    md_download = client.get(markdown.json()["download_url"], headers=auth_headers)
    assert md_download.status_code == 200
    assert "# pytest 导出分享 仿真报告" in md_download.text
    assert "**突出防水和续航**" in md_download.text

    alias = client.post(
        f"/api/simulations/{project_id}/export",
        headers=auth_headers,
        json={"format": "json"},
    )
    assert alias.status_code == 201
    assert alias.json()["download_url"].startswith("/api/exports/")
    assert "## 图表数据摘要" in md_download.text
    assert "测试智能手机" in md_download.text
    assert "sk-should-not-appear" not in md_download.text


def test_export_rejects_unfinished_and_can_create_pdf(
    client: TestClient,
    auth_headers: dict[str, str],
    monkeypatch,
) -> None:
    created = client.post("/api/simulations", headers=auth_headers, json={"project_name": "pytest 未完成"})
    created.raise_for_status()
    project_id = created.json()["id"]

    unfinished = client.post(
        f"/api/simulations/{project_id}/exports",
        headers=auth_headers,
        json={"format": "json"},
    )
    assert unfinished.status_code == 409

    _promote_to_pro(client, auth_headers)
    completed_id = _create_completed_project(client, auth_headers)

    def fake_write_export_file(task, project) -> Any:
        path = export_file_path(task)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"%PDF-1.4\n%pytest\n")
        task.status = "completed"
        task.download_url = f"/api/exports/{task.id}/download"
        task.completed_at = utc_now_naive()
        return path

    monkeypatch.setattr(export_worker, "write_export_file", fake_write_export_file)
    pdf = client.post(
        f"/api/simulations/{completed_id}/exports",
        headers=auth_headers,
        json={"format": "pdf"},
    )
    assert pdf.status_code == 201
    assert pdf.json()["format"] == "pdf"
    assert pdf.json()["status"] == "queued"

    assert export_worker.run_loop(once=True, timeout=1) == 0
    status_response = client.get(f"/api/exports/{pdf.json()['export_task_id']}", headers=auth_headers)
    assert status_response.status_code == 200
    assert status_response.json()["status"] == "completed"

    downloaded = client.get(status_response.json()["download_url"], headers=auth_headers)
    assert downloaded.status_code == 200
    assert downloaded.content.startswith(b"%PDF")


def test_basic_user_cannot_export(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    project_id = _create_completed_project(client, auth_headers)
    response = client.post(
        f"/api/simulations/{project_id}/exports",
        headers=auth_headers,
        json={"format": "json"},
    )
    assert response.status_code == 403
    assert response.json()["code"] == "EXPORT_FORBIDDEN"


def test_pro_user_can_export_excel(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    _promote_to_pro(client, auth_headers)
    report = _sample_report()
    report["strategy_recommendations"] = [
        {
            "strategy": "强化耐用性宣传",
            "actions": ["展示耐用测试", "提供延长保修"],
            "expected_impact": "缓解售后顾虑",
        }
    ]
    project_id = _create_completed_project(client, auth_headers, report)
    response = client.post(
        f"/api/simulations/{project_id}/exports",
        headers=auth_headers,
        json={"format": "excel"},
    )
    assert response.status_code == 201
    assert response.json()["format"] == "excel"
    downloaded = client.get(response.json()["download_url"], headers=auth_headers)
    assert downloaded.status_code == 200
    assert downloaded.content[:2] == b"PK"
    workbook = load_workbook(BytesIO(downloaded.content), read_only=True)
    assert "目标人群画像" in workbook.sheetnames
    assert "策略建议" in workbook.sheetnames
    assert "图表_市场份额" in workbook.sheetnames
    assert "图表_竞品雷达" in workbook.sheetnames
    assert "图表_社交演化" in workbook.sheetnames
    assert "社交传播轮次" in workbook.sheetnames
    strategy_sheet = workbook["策略建议"]
    assert strategy_sheet.cell(row=1, column=1).value == "策略"
    assert strategy_sheet.cell(row=2, column=1).value == "强化耐用性宣传"
    assert strategy_sheet.cell(row=2, column=2).value == "展示耐用测试；提供延长保修"


def test_excel_export_cleans_mixed_chart_rows(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    _promote_to_pro(client, auth_headers)
    report = _sample_report()
    report["chart_data"]["channel_effect"] = ["短视频渠道待确认", ["线下体验", 0.35], None]
    report["evidence_used"] = ["用户关注续航", {"source": "用户画像数据_pytest", "snippet": "关注价格"}]
    project_id = _create_completed_project(client, auth_headers, report)

    response = client.post(
        f"/api/simulations/{project_id}/exports",
        headers=auth_headers,
        json={"format": "excel"},
    )

    assert response.status_code == 201
    downloaded = client.get(response.json()["download_url"], headers=auth_headers)
    assert downloaded.status_code == 200
    workbook = load_workbook(BytesIO(downloaded.content), read_only=True)
    sheet = workbook["图表_渠道贡献"]
    assert sheet.cell(row=1, column=1).value == "value"
    assert sheet.cell(row=2, column=1).value == "短视频渠道待确认"


def test_share_token_public_report_is_sanitized_and_can_be_disabled(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    _promote_to_pro(client, auth_headers)
    project_id = _create_completed_project(client, auth_headers)

    created = client.post(
        f"/api/simulations/{project_id}/share-tokens",
        headers=auth_headers,
        json={"expires_in_hours": 72},
    )
    assert created.status_code == 201
    body = created.json()
    assert body["token"]
    assert "/share/" in body["share_url"]
    assert body["share_url"] == body["frontend_share_url"]
    assert "/api/share/" in body["api_share_url"]

    public = client.get(f"/api/share/{body['token']}")
    assert public.status_code == 200
    public_text = public.text
    assert "pytest 完整报告" in public_text
    assert "chart_data" in public_text
    assert "prompt_trace" not in public_text
    assert "formal_test_log_path" not in public_text
    assert "sk-should-not-appear" not in public_text
    assert '"raw"' not in public_text
    assert "user_id" not in public_text
    assert "project_id" not in public.json()
    assert "snapshot_hash" not in public.json()
    assert "snapshot_id" not in public.text
    assert "share_token_id" not in public.json()

    disabled = client.delete(f"/api/share-tokens/{body['id']}", headers=auth_headers)
    assert disabled.status_code == 200
    assert disabled.json()["is_active"] is False
    assert client.get(f"/api/share/{body['token']}").status_code == 404

    alias_created = client.post(
        f"/api/simulations/{project_id}/share",
        headers=auth_headers,
        json={"expires_in_hours": 72},
    )
    assert alias_created.status_code == 201
    revoked = client.post(f"/api/share/{alias_created.json()['token']}/revoke", headers=auth_headers)
    assert revoked.status_code == 200
    assert revoked.json()["is_active"] is False


def test_basic_user_cannot_share(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    project_id = _create_completed_project(client, auth_headers)
    response = client.post(
        f"/api/simulations/{project_id}/share-tokens",
        headers=auth_headers,
        json={"expires_in_hours": 72},
    )
    assert response.status_code == 403
    assert response.json()["code"] == "SHARE_FORBIDDEN"


def test_expired_share_token_is_not_accessible(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    _promote_to_pro(client, auth_headers)
    project_id = _create_completed_project(client, auth_headers)
    created = client.post(
        f"/api/simulations/{project_id}/share-tokens",
        headers=auth_headers,
        json={"expires_in_hours": 1},
    )
    created.raise_for_status()
    token = created.json()["token"]
    token_id = created.json()["id"]

    with SessionLocal() as db:
        item = db.get(ShareToken, token_id)
        assert item is not None
        item.expires_at = utc_now_naive() - timedelta(minutes=1)
        db.commit()

    assert client.get(f"/api/share/{token}").status_code == 404


def test_distill_client_is_disabled_by_default() -> None:
    result = run_distill_checks_if_enabled({}, [], [])
    assert result["enabled"] is False
    assert result["status"] == "disabled"
    assert result["checked_samples"] == 0
    assert result["warning_level"] == "info"


def test_debug_distill_endpoint_uses_standard_shape(client: TestClient) -> None:
    response = client.post("/api/debug/distill/check", json={"snapshot": {}, "agents": [], "purchase_decisions": []})
    assert response.status_code == 200
    assert response.json()["status"] == "disabled"
    assert "validation_batch_id" in response.json()


def test_pdf_preflight_reports_frontend_unreachable(monkeypatch) -> None:
    class DummyError(RuntimeError):
        pass

    def fake_get(*args, **kwargs):
        raise DummyError("frontend down")

    monkeypatch.setattr(export_service.settings, "playwright_browsers_path", str(Path("logs/test_runs/playwright-browsers")))
    monkeypatch.setattr(export_service.settings, "frontend_base_url", "http://127.0.0.1:59999")
    monkeypatch.setattr(export_service.httpx, "get", fake_get)

    result = export_service.check_pdf_render_prerequisites()

    assert result["ok"] is False
    assert result["checks"]["frontend"]["ok"] is False
    assert "前端服务不可访问" in result["checks"]["frontend"]["message"]
