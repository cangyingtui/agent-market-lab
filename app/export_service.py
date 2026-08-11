from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any
from datetime import timedelta

import httpx
from fastapi import HTTPException
from jose import JWTError, jwt

from app.config import settings
from app.crowd_profile import PROFILE_LABELS, normalize_crowd_profile, normalize_crowd_segments, normalize_profile
from app.models import ExportTask, SimulationProject
from app.strategy_recommendations import normalize_strategy_recommendations, strategy_recommendation_rows
from app.time_utils import utc_now_iso, utc_now_naive


PRIVATE_STRIP_KEYS = {
    "prompt_trace",
    "formal_test_log_path",
    "llm_raw_response",
    "raw_response",
    "raw_response_truncated",
    "api_key",
    "authorization",
    "headers",
}
PUBLIC_STRIP_KEYS = PRIVATE_STRIP_KEYS | {
    "raw",
    "queries",
    "llm_error",
    "fallback_reason",
    "snapshot_hash",
    "snapshot_id",
    "neighbors",
    "base_maut_scores",
    "neighbor_purchase_intent_avg",
    "social_score_change",
    "economics",
    "strategy_economics",
    "gross_margin_pct",
    "unit_promotion_cost_cny",
    "total_budget_cny",
    "gross_profit_per_unit",
    "promotion_burden_per_unit",
    "contribution_after_promotion",
    "margin_safety_pct",
}
BUSINESS_STRIP_KEYS = {"distill_summary"}
WEB_REPORT_OMIT_KEYS = {
    "formatted_evidence",
    "final_rag_evidence",
    "purchase_decisions",
    "queries",
}
WEB_REPORT_LIST_LIMITS = {
    "agent_samples": 12,
    "data_enrichment_candidates": 20,
    "evidence_used": 40,
    "market_strategy_evidence": 20,
    "structured_product_evidence": 40,
    "user_profile_evidence": 20,
}
WEB_RAG_EVIDENCE_GROUP_LIMIT = 20


def _with_evidence_confidence(report: dict[str, Any]) -> dict[str, Any]:
    aggregation = report.get("aggregation") if isinstance(report.get("aggregation"), dict) else {}
    confidence = aggregation.get("confidence") if isinstance(aggregation.get("confidence"), dict) else {}
    if isinstance(confidence.get("components"), dict):
        return report
    agents = report.get("agent_samples")
    decisions = report.get("purchase_decisions")
    evidence = report.get("rag_evidence")
    if not isinstance(agents, list) or not isinstance(decisions, list) or not isinstance(evidence, dict):
        return report
    try:
        from engine.aggregation import aggregate_results

        recomputed = aggregate_results(
            [item for item in agents if isinstance(item, dict)],
            [item for item in decisions if isinstance(item, dict)],
            evidence,
        )
    except Exception:
        return report
    copied = dict(report)
    upgraded_aggregation = dict(aggregation)
    upgraded_aggregation["confidence"] = recomputed.get("confidence", {})
    upgraded_aggregation.setdefault("evidence_quality", recomputed.get("evidence_quality", {}))
    upgraded_aggregation.setdefault("rag_evidence_quality", recomputed.get("rag_evidence_quality", {}))
    upgraded_aggregation.setdefault("crowd_profile_quality", recomputed.get("crowd_profile_quality", {}))
    copied["aggregation"] = upgraded_aggregation
    copied.setdefault("data_quality", upgraded_aggregation.get("evidence_quality", {}))
    return copied


def _sanitize_value(value: Any, *, public: bool) -> Any:
    strip_keys = PUBLIC_STRIP_KEYS if public else PRIVATE_STRIP_KEYS
    if isinstance(value, dict):
        sanitized: dict[str, Any] = {}
        for key, item in value.items():
            key_text = str(key).lower()
            if key_text in strip_keys or key_text in BUSINESS_STRIP_KEYS or "api_key" in key_text or "secret" in key_text:
                continue
            sanitized[str(key)] = _sanitize_value(item, public=public)
        return sanitized
    if isinstance(value, list):
        return [_sanitize_value(item, public=public) for item in value]
    return value


def sanitize_report(report: dict[str, Any] | None, *, public: bool = False) -> dict[str, Any]:
    if not isinstance(report, dict):
        return {}
    sanitized = _sanitize_value(_with_evidence_confidence(report), public=public)
    if public:
        sanitized["public_view"] = True
    return sanitized


def _compact_web_value(value: Any, *, key: str | None = None) -> Any:
    if key in WEB_REPORT_OMIT_KEYS:
        return None
    if isinstance(value, dict):
        compacted: dict[str, Any] = {}
        for child_key, child_value in value.items():
            child_key_text = str(child_key)
            if child_key_text in WEB_REPORT_OMIT_KEYS:
                continue
            compacted[child_key_text] = _compact_web_value(
                child_value,
                key="rag_evidence" if key == "rag_evidence" else child_key_text,
            )
        return compacted
    if isinstance(value, list):
        limit = WEB_REPORT_LIST_LIMITS.get(key or "")
        if key == "rag_evidence":
            limit = WEB_RAG_EVIDENCE_GROUP_LIMIT
        items = value[:limit] if limit is not None else value
        return [_compact_web_value(item) for item in items]
    return value


def sanitize_web_report(report: dict[str, Any] | None, *, public: bool = False) -> dict[str, Any]:
    sanitized = sanitize_report(report, public=public)
    # Browser views only need summaries and chart-ready rows. Keep full detail in
    # result_data for JSON, Markdown, Excel and PDF export generation.
    # Owner-facing reports may display the optional economics supplied by that
    # owner. Only public share/print views use the stricter public stripping set.
    web_safe = _sanitize_value(sanitized, public=public)
    compacted = _compact_web_value(web_safe)
    return compacted if isinstance(compacted, dict) else {}


def _safe_positive_price(value: Any) -> float | None:
    try:
        parsed = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def _project_price_for_export(project: SimulationProject, report: dict[str, Any]) -> float | None:
    snapshot = project.config_snapshot if isinstance(project.config_snapshot, dict) else {}
    snapshot_product = snapshot.get("product_definition") if isinstance(snapshot.get("product_definition"), dict) else {}
    report_product = report.get("product_definition") if isinstance(report.get("product_definition"), dict) else {}
    result_data = project.result_data if isinstance(project.result_data, dict) else {}
    result_product = result_data.get("product_definition") if isinstance(result_data.get("product_definition"), dict) else {}
    project_product = project.product_definition if isinstance(project.product_definition, dict) else {}
    pricing = report.get("pricing_analysis") if isinstance(report.get("pricing_analysis"), dict) else {}
    for value in (
        snapshot_product.get("price_cny"),
        report_product.get("price_cny"),
        result_product.get("price_cny"),
        project_product.get("price_cny"),
        pricing.get("reference_price"),
    ):
        parsed = _safe_positive_price(value)
        if parsed is not None:
            return parsed
    return None


def with_project_report_fallbacks(
    report: dict[str, Any],
    project: SimulationProject,
    *,
    include_market_config: bool = True,
) -> dict[str, Any]:
    enriched = dict(report)
    market_config = project.market_config if isinstance(project.market_config, dict) else {}
    product_price = _project_price_for_export(project, enriched)
    if product_price is not None:
        product_definition = dict(enriched.get("product_definition") if isinstance(enriched.get("product_definition"), dict) else {})
        product_definition.setdefault("price_cny", product_price)
        enriched["product_definition"] = product_definition
        pricing = dict(enriched.get("pricing_analysis") if isinstance(enriched.get("pricing_analysis"), dict) else {})
        pricing.setdefault("reference_price", product_price)
        enriched["pricing_analysis"] = pricing
    if market_config:
        if include_market_config:
            enriched.setdefault("market_config", market_config)
        if not enriched.get("crowd_segments"):
            enriched["crowd_segments"] = normalize_crowd_segments(market_config)
        if not enriched.get("crowd_profile"):
            enriched["crowd_profile"] = normalize_crowd_profile(market_config)
        if not enriched.get("selected_strategies"):
            enriched["selected_strategies"] = selected_strategy_rows(enriched, market_config)
    return enriched


def build_report_payload(project: SimulationProject, *, public: bool = False, compact: bool = False) -> dict[str, Any]:
    report = sanitize_web_report(project.result_data, public=public) if compact else sanitize_report(project.result_data, public=public)
    report = with_project_report_fallbacks(report, project, include_market_config=not public)
    payload = {
        "project_name": project.project_name,
        "status": project.status,
        "plan_type_used": project.plan_type_used or "basic",
        "generated_at": utc_now_iso(),
        "report": report,
    }
    if not public:
        payload.update({"project_id": project.id, "snapshot_hash": project.snapshot_hash})
    return payload


def _normalize_table_row(row: Any) -> dict[str, Any]:
    if isinstance(row, dict):
        return row
    if row is None:
        return {"value": "暂无数据"}
    if isinstance(row, list):
        return {"value": json.dumps(row, ensure_ascii=False, default=str)}
    return {"value": row}


def _normalize_table_rows(rows: Any) -> list[dict[str, Any]]:
    if isinstance(rows, list):
        return [_normalize_table_row(row) for row in rows]
    if isinstance(rows, dict):
        return _key_value_rows(rows)
    if rows in (None, ""):
        return []
    return [_normalize_table_row(rows)]


def _readable_markdown_value(value: Any, max_length: int = 360) -> str:
    if value in (None, ""):
        return "-"
    if isinstance(value, list):
        text = "、".join(_readable_markdown_value(item, max_length=120) for item in value if item not in (None, ""))
    elif isinstance(value, dict):
        preferred_keys = (
            "insight",
            "summary",
            "description",
            "name",
            "product_name",
            "strategy",
            "expected_impact",
            "source",
        )
        parts = [str(value.get(key)).strip() for key in preferred_keys if value.get(key) not in (None, "")]
        if not parts:
            parts = [f"{key}: {_readable_markdown_value(item, max_length=80)}" for key, item in value.items() if item not in (None, "")]
        text = "；".join(part for part in parts if part)
    else:
        text = str(value).strip()
    text = " ".join(text.replace("\r", " ").replace("\n", " ").split())
    if len(text) > max_length:
        return text[: max_length - 1] + "…"
    return text or "-"


def _markdown_cell(value: Any) -> str:
    return _readable_markdown_value(value).replace("|", "\\|")


MARKDOWN_HEADER_LABELS = {
    "name": "名称",
    "role": "类型",
    "share": "占比",
    "source": "来源",
    "source_type": "来源类型",
    "snippet": "片段",
    "summary": "摘要",
    "insight": "洞察",
    "content": "内容",
    "detail": "详情",
    "value": "数值",
    "count": "样本数",
    "importance": "重要性",
    "weight": "权重",
    "price": "价格",
    "intent": "购买意愿",
    "round": "轮次",
    "social_influence": "社会影响",
    "max_score_change": "最大变化",
    "roi": "ROI",
    "intensity": "投放强度",
    "discount": "折扣",
}


def _markdown_table(rows: list[dict[str, Any]] | Any, preferred_keys: list[str] | None = None) -> list[str]:
    rows = _normalize_table_rows(rows)
    if not rows:
        return ["暂无"]
    keys: list[str] = []
    for key in preferred_keys or []:
        if any(key in row for row in rows):
            keys.append(key)
    for row in rows:
        for key in row.keys():
            if key not in keys and not isinstance(row.get(key), (dict, list)):
                keys.append(key)
    if not keys:
        return ["暂无"]
    headers = [MARKDOWN_HEADER_LABELS.get(key, key) for key in keys]
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(keys)) + " |"]
    for row in rows:
        values = []
        for key in keys:
            values.append(_markdown_cell(row.get(key, "")))
        lines.append("| " + " | ".join(values) + " |")
    return lines


def _key_value_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [{"指标": key, "值": value} for key, value in data.items()]


def _strategy_name(value: Any) -> str:
    if isinstance(value, dict):
        for key in ("name", "strategy", "策略", "title", "label"):
            raw = value.get(key)
            if raw:
                return str(raw).strip()
        return ""
    if value is None:
        return ""
    return str(value).strip()


def selected_strategy_rows(report: dict[str, Any], market_config: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    raw = report.get("selected_strategies")
    report_market = report.get("market_config") if isinstance(report.get("market_config"), dict) else {}
    source_market = market_config if isinstance(market_config, dict) and market_config else report_market
    strategy_details = {}
    if isinstance(report.get("strategy_details"), dict):
        strategy_details.update(report["strategy_details"])
    if isinstance(source_market.get("strategy_details"), dict):
        strategy_details.update(source_market["strategy_details"])
    if not isinstance(raw, list) or not raw:
        raw = source_market.get("strategies") if isinstance(source_market.get("strategies"), list) else []
        if not raw and source_market.get("strategy"):
            raw = [source_market.get("strategy")]
    rows: list[dict[str, Any]] = []
    for index, item in enumerate(raw, 1):
        name = _strategy_name(item)
        if not name:
            continue
        row = {"序号": index, "策略": name}
        detail = strategy_details.get(name) if isinstance(strategy_details.get(name), dict) else {}
        source_item = {**detail, **item} if isinstance(item, dict) else detail
        channels = source_item.get("channels") or source_item.get("touch_channels")
        if isinstance(channels, list):
            row["渠道"] = "、".join(str(value) for value in channels if value)
        elif channels:
            row["渠道"] = str(channels)
        for key, label in (
            ("description", "说明"),
            ("core_selling_point", "核心卖点"),
            ("target_segment", "目标客群"),
            ("benefit", "优惠/权益"),
            ("action", "执行动作"),
            ("budget_intensity", "预算强度"),
            ("risk_note", "风险说明"),
            ("intensity", "强度"),
            ("price_discount", "折扣"),
        ):
            if source_item.get(key) not in (None, ""):
                row[label] = source_item.get(key)
        economics = source_item.get("economics") if isinstance(source_item.get("economics"), dict) else {}
        for key, label in (
            ("gross_margin_pct", "基础毛利率"),
            ("discount_pct", "让利比例"),
            ("unit_promotion_cost_cny", "单笔推广成本"),
            ("total_budget_cny", "总预算"),
        ):
            if economics.get(key) not in (None, ""):
                row[label] = economics.get(key)
        rows.append(row)
    return rows


def crowd_profile_rows(report: dict[str, Any], market_config: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    segments = normalize_crowd_segments({"crowd_segments": report.get("crowd_segments")})
    if not segments and isinstance(report.get("market_config"), dict):
        segments = normalize_crowd_segments(report.get("market_config"))
    if not segments and market_config:
        segments = normalize_crowd_segments(market_config)
    if not segments:
        target_segments = report.get("target_segments") if isinstance(report.get("target_segments"), list) else []
        segments = [
            {
                "name": item.get("name") or "目标用户",
                "ratio": item.get("ratio") or 100,
                "profile": normalize_profile(item.get("crowd_profile"), str(item.get("name") or "目标用户")),
            }
            for item in target_segments
            if isinstance(item, dict)
        ]
    if not segments:
        profile = report.get("crowd_profile") if isinstance(report.get("crowd_profile"), dict) else {}
        if not profile:
            profile_source = market_config if market_config else {"target_crowd": report.get("target_crowd"), "crowd_profile": report.get("crowd_profile")}
            profile = normalize_crowd_profile(profile_source)
        segments = [{"name": profile.get("name") or "目标用户", "ratio": 100, "profile": profile}]
    rows: list[dict[str, Any]] = []
    for segment in segments:
        profile = normalize_profile(segment.get("profile"), str(segment.get("name") or "目标用户"))
        for key, label in {"name": "目标人群", **PROFILE_LABELS}.items():
            value = profile.get(key)
            if isinstance(value, list):
                value = "、".join(str(item) for item in value if item)
            if value:
                rows.append(
                    {
                        "客群": segment.get("name") or "目标用户",
                        "占比": f"{segment.get('ratio') or 100}%",
                        "字段": label,
                        "内容": value,
                    }
                )
    return rows


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _pct(value: Any, *, scale: float = 1.0) -> str:
    return f"{_as_float(value) * scale:.1f}%"


def _status_label(value: Any) -> str:
    return {
        "draft": "未提交",
        "submitted": "已提交",
        "queued": "等待生成",
        "running": "生成中",
        "report_waiting": "报告生成中",
        "completed": "已完成",
        "failed": "生成中断",
        "cancelled": "已取消",
    }.get(str(value or "").strip(), str(value or "-"))


def _plan_label(value: Any) -> str:
    return "专业版" if str(value or "basic") == "pro" else "普通版"


def _profile_summary(profile: dict[str, Any]) -> str:
    parts: list[str] = []
    for key, label in PROFILE_LABELS.items():
        value = profile.get(key)
        if isinstance(value, list):
            value = "、".join(str(item) for item in value if item)
        if value not in (None, ""):
            parts.append(f"{label}：{value}")
    return "；".join(parts[:8]) or "-"


def target_segment_summary_rows(report: dict[str, Any], market_config: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    segments = normalize_crowd_segments({"crowd_segments": report.get("crowd_segments")})
    if not segments and isinstance(report.get("market_config"), dict):
        segments = normalize_crowd_segments(report.get("market_config"))
    if not segments and market_config:
        segments = normalize_crowd_segments(market_config)
    if not segments and isinstance(report.get("target_segments"), list):
        segments = [
            {
                "name": item.get("name") or "目标用户",
                "ratio": item.get("ratio") or 100,
                "insight": item.get("insight"),
                "profile": item.get("profile") or item.get("crowd_profile") or {},
            }
            for item in report.get("target_segments") or []
            if isinstance(item, dict)
        ]
    if not segments:
        return []
    rows: list[dict[str, Any]] = []
    for segment in segments:
        profile = normalize_profile(segment.get("profile"), str(segment.get("name") or "目标用户"))
        rows.append(
            {
                "客群": segment.get("name") or profile.get("name") or "目标用户",
                "占比": f"{segment.get('ratio') or 100}%",
                "画像摘要": segment.get("insight") or _profile_summary(profile),
            }
        )
    return rows


def competitor_insight_rows(report: dict[str, Any], limit: int = 12) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, item in enumerate(report.get("competitor_insights") or [], 1):
        if not isinstance(item, dict):
            rows.append({"序号": index, "来源": "-", "评分": "-", "洞察": item})
            continue
        rows.append(
            {
                "序号": index,
                "来源": item.get("source") or item.get("product_name") or item.get("name") or "-",
                "评分": round(_as_float(item.get("score")), 2) if item.get("score") not in (None, "") else "-",
                "洞察": item.get("insight") or item.get("summary") or item,
            }
        )
        if len(rows) >= limit:
            break
    return rows


def evidence_summary_rows(report: dict[str, Any], limit: int = 20) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, item in enumerate(report.get("evidence_used") or [], 1):
        if not isinstance(item, dict):
            rows.append({"序号": index, "来源": "-", "评分": "-", "摘要": item})
            continue
        rows.append(
            {
                "序号": index,
                "来源": item.get("source") or item.get("source_type") or item.get("group") or "-",
                "评分": round(_as_float(item.get("score")), 2) if item.get("score") not in (None, "") else "-",
                "摘要": item.get("insight") or item.get("content") or item.get("summary") or item,
            }
        )
        if len(rows) >= limit:
            break
    return rows


def pricing_analysis_rows(report: dict[str, Any]) -> list[dict[str, Any]]:
    pricing = report.get("pricing_analysis")
    if not isinstance(pricing, dict):
        return []
    preferred = [
        ("summary", "结论"),
        ("recommended_price", "建议价格"),
        ("price_band", "建议价格带"),
        ("price_sensitivity", "价格敏感性"),
        ("competitor_price_coverage", "竞品价格覆盖"),
    ]
    rows = []
    used = set()
    for key, label in preferred:
        if key in pricing:
            rows.append({"项目": label, "内容": pricing.get(key)})
            used.add(key)
    for key, value in pricing.items():
        if key not in used:
            rows.append({"项目": key, "内容": value})
    return rows


def _decision_dimension_rows(report: dict[str, Any]) -> list[dict[str, Any]]:
    aggregation = report.get("aggregation") if isinstance(report.get("aggregation"), dict) else {}
    decision_model = report.get("decision_model") if isinstance(report.get("decision_model"), dict) else {}
    dimensions = aggregation.get("dimension_scores") or decision_model.get("dimension_scores") or {}
    if not isinstance(dimensions, dict):
        return []
    rows: list[dict[str, Any]] = []
    for key, value in dimensions.items():
        if not isinstance(value, dict):
            continue
        rows.append(
            {
                "维度": value.get("label") or key,
                "均值": _pct(value.get("avg_score"), scale=100),
                "加权贡献": _pct(value.get("weighted_contribution"), scale=100),
                "权重": _as_float(value.get("weight")),
            }
        )
    return rows


def build_business_analysis_lines(report: dict[str, Any]) -> list[str]:
    chart_data = report.get("chart_data") if isinstance(report.get("chart_data"), dict) else {}
    overview = chart_data.get("overview_metrics") if isinstance(chart_data.get("overview_metrics"), dict) else {}
    aggregation = report.get("aggregation") if isinstance(report.get("aggregation"), dict) else {}
    confidence = aggregation.get("confidence") if isinstance(aggregation.get("confidence"), dict) else {}
    dimensions = _decision_dimension_rows(report)
    sorted_dimensions = sorted(
        dimensions,
        key=lambda row: _as_float(str(row.get("加权贡献", "0")).replace("%", "")),
        reverse=True,
    )
    top = sorted_dimensions[0] if sorted_dimensions else {"维度": "功能匹配度", "加权贡献": "0.0%"}
    weak = sorted_dimensions[-1] if sorted_dimensions else {"维度": "价格接受度", "加权贡献": "0.0%"}
    price_quality = aggregation.get("evidence_quality") if isinstance(aggregation.get("evidence_quality"), dict) else report.get("data_quality") or {}
    if not isinstance(price_quality, dict):
        price_quality = {}
    market_scope = report.get("market_share_scope") if isinstance(report.get("market_share_scope"), dict) else chart_data.get("market_share_scope") or {}
    scenario_share = market_scope.get("full_market_scenario_share")
    rci = market_scope.get("relative_competitiveness_index")
    market_line = (
        f"购买意愿指数为 {_pct(overview.get('purchase_intent_index'))}，仿真环境份额为 {_pct(overview.get('estimated_market_share'))}；"
        "该份额是本品与本轮已选竞品的封闭归一化结果，不代表真实全市场预测。"
    )
    if scenario_share is not None:
        market_line += f" 按 {market_scope.get('assumed_market_competitor_count') or '-'} 个竞品换算的全市场情景份额为 {_pct(scenario_share)}，RCI 为 {_as_float(rci):.2f}。"
    return [
        market_line,
        f"五维拆解中，{top.get('维度')} 是当前主要支撑项，加权贡献约 {top.get('加权贡献')}；{weak.get('维度')} 是相对薄弱项，建议在销售表达或产品配置上补强。",
        f"当前证据置信度为 {confidence.get('label') or '-'}（{_pct(confidence.get('score'), scale=100)}），竞品价格覆盖率为 {_pct(price_quality.get('price_coverage_pct'))}。若价格覆盖率偏低，定价结论应视为方向性建议。",
    ]


def render_markdown_report(project: SimulationProject) -> str:
    report = sanitize_report(project.result_data, public=False)
    chart_data = report.get("chart_data") if isinstance(report.get("chart_data"), dict) else {}
    market_config = project.market_config if isinstance(project.market_config, dict) else {}
    lines = [
        f"# {project.project_name} 仿真报告",
        "",
        f"- 仿真编号：SIM-{str(project.id).zfill(5)}",
        f"- 状态：{_status_label(project.status)}",
        f"- 项目版本：{_plan_label(project.plan_type_used)}",
        f"- 导出时间：{utc_now_iso()}",
        "",
        "## 执行摘要",
        "",
        str(report.get("executive_summary") or "暂无摘要"),
        "",
        "## 业务解读",
        "",
        *[f"- {line}" for line in build_business_analysis_lines(report)],
        "",
        "## 购买模型五维拆解",
        "",
        *_markdown_table(_decision_dimension_rows(report), ["维度", "均值", "加权贡献", "权重"]),
        "",
        "## 图表数据摘要",
        "",
        "### 概览指标",
        "",
        *_markdown_table(_key_value_rows(chart_data.get("overview_metrics") or {}), ["指标", "值"]),
        "",
        "### 仿真环境份额",
        "",
        *_markdown_table(chart_data.get("market_share") if isinstance(chart_data.get("market_share"), list) else [], ["name", "role", "share", "source"]),
        "",
        "### 市场份额口径",
        "",
        *_markdown_table(_key_value_rows(chart_data.get("market_share_scope") or report.get("market_share_scope") or {}), ["指标", "值"]),
        "",
        "### 全市场情景换算",
        "",
        *_markdown_table(chart_data.get("market_share_scenarios") if isinstance(chart_data.get("market_share_scenarios"), list) else [], ["competitor_count", "share"]),
        "",
        "### 购买意愿",
        "",
        *_markdown_table(
            chart_data.get("purchase_intent_by_segment") if isinstance(chart_data.get("purchase_intent_by_segment"), list) else [],
            ["name", "value", "count"],
        ),
        "",
        "### 功能重要性",
        "",
        *_markdown_table(chart_data.get("param_importance") if isinstance(chart_data.get("param_importance"), list) else [], ["name", "importance", "weight", "value"]),
        "",
        "### 价格敏感曲线",
        "",
        *_markdown_table(chart_data.get("price_sensitivity") if isinstance(chart_data.get("price_sensitivity"), list) else [], ["price", "intent"]),
        "",
        "### 仿真 ROI",
        "",
        *(
            [str((chart_data.get("simulation_boundaries") or report.get("simulation_boundaries") or {}).get("expert_strategy_note") or ""), "", str((chart_data.get("simulation_boundaries") or report.get("simulation_boundaries") or {}).get("simulation_roi_note") or ""), ""]
            if chart_data.get("commercial_model_version") or report.get("commercial_model_version")
            else []
        ),
        *_markdown_table(chart_data.get("strategy_roi") if isinstance(chart_data.get("strategy_roi"), list) else [], ["name", "roi", "recommendation_priority", "reach_score", "conversion_lift", "cost_pressure", "risk_penalty", "margin_safety_pct", "expert_basis"]),
        "",
        "### 渠道贡献拆解",
        "",
        *_markdown_table(chart_data.get("channel_effect") if isinstance(chart_data.get("channel_effect"), list) else [], ["name", "share", "reach_score", "conversion_score", "acquisition_cost_pressure", "crowd_match", "scene_match"]),
        "",
        "## 社交传播分析",
        "",
        *_markdown_table(
            _key_value_rows(
                {
                    key: (report.get("social_simulation") or {}).get(key)
                    for key in ("rounds_executed", "converged", "node_count", "edge_count", "average_degree", "topology")
                }
            ),
            ["指标", "值"],
        ),
        "",
        "### 社交演化",
        "",
        *_markdown_table(chart_data.get("social_evolution") if isinstance(chart_data.get("social_evolution"), list) else [], ["round", "name", "value", "social_influence", "max_score_change"]),
        "",
        "## 目标人群",
        "",
        *_markdown_table(crowd_profile_rows(report, market_config), ["客群", "占比", "字段", "内容"]),
        "",
        "### 目标人群摘要",
        "",
    ]
    lines.extend(_markdown_table(target_segment_summary_rows(report, market_config), ["客群", "占比", "画像摘要"]))

    lines.extend(["", "## 竞品洞察", ""])
    lines.extend(_markdown_table(competitor_insight_rows(report), ["序号", "来源", "评分", "洞察"]))

    lines.extend(["", "## 已配置营销策略", ""])
    lines.extend(_markdown_table(selected_strategy_rows(report, market_config), ["序号", "策略", "渠道", "说明"]))

    lines.extend(["", "## 定价分析", ""])
    lines.extend(_markdown_table(pricing_analysis_rows(report), ["项目", "内容"]))
    lines.extend(["", "## 策略建议", ""])
    recommendations = normalize_strategy_recommendations(report.get("strategy_recommendations"))
    for item in recommendations:
        lines.append(f"- **{item['strategy']}**")
        if item["actions"]:
            lines.append(f"  - 执行动作：{'；'.join(item['actions'])}")
        if item["expected_impact"]:
            lines.append(f"  - 预期影响：{item['expected_impact']}")
        if item.get("recommendation_priority"):
            lines.append(f"  - 推荐优先级：{item['recommendation_priority']}")
        if item.get("expert_basis"):
            lines.append(f"  - 专家依据：{item['expert_basis']}")
        if item.get("applicable_conditions"):
            lines.append(f"  - 适用条件：{'；'.join(item['applicable_conditions'])}")
        if item.get("cost_risk"):
            lines.append(f"  - 成本风险：{item['cost_risk']}")
    if not recommendations:
        lines.append("- 暂无")

    lines.extend(["", "## 风险提醒", ""])
    for item in report.get("risk_warnings") or []:
        lines.append(f"- {item}")
    if not report.get("risk_warnings"):
        lines.append("- 暂无")

    lines.extend(["", "## 证据摘要", ""])
    lines.extend(_markdown_table(evidence_summary_rows(report), ["序号", "来源", "评分", "摘要"]))

    lines.append("")
    return "\n".join(lines)


def export_file_path(task: ExportTask) -> Path:
    suffix = {"markdown": "md", "excel": "xlsx", "pdf": "pdf"}.get(task.format, task.format)
    return settings.resolve_path(settings.export_dir) / f"project_{task.project_id}" / f"export_{task.id}.{suffix}"


def create_pdf_render_token(project_id: int, export_task_id: int) -> str:
    expires_at = utc_now_naive() + timedelta(minutes=10)
    return jwt.encode(
        {
            "sub": "pdf_render",
            "project_id": project_id,
            "export_task_id": export_task_id,
            "exp": expires_at,
        },
        settings.jwt_secret_key,
        algorithm=settings.jwt_algorithm,
    )


def decode_pdf_render_token(token: str) -> dict[str, Any]:
    try:
        payload = jwt.decode(token, settings.jwt_secret_key, algorithms=[settings.jwt_algorithm])
    except JWTError as exc:
        raise HTTPException(status_code=404, detail="PDF 渲染令牌无效或已过期") from exc
    if payload.get("sub") != "pdf_render":
        raise HTTPException(status_code=404, detail="PDF 渲染令牌无效")
    return payload


def configure_playwright_browsers_path() -> Path:
    browsers_path = settings.resolve_path(settings.playwright_browsers_path)
    browsers_path.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", str(browsers_path))
    return browsers_path


def check_pdf_render_prerequisites() -> dict[str, Any]:
    browsers_path = configure_playwright_browsers_path()
    result: dict[str, Any] = {
        "ok": True,
        "frontend_base_url": settings.frontend_base_url,
        "browsers_path": str(browsers_path),
        "checks": {},
    }
    try:
        from playwright.sync_api import sync_playwright  # noqa: F401

        result["checks"]["playwright"] = {"ok": True}
    except Exception as exc:
        result["ok"] = False
        result["checks"]["playwright"] = {
            "ok": False,
            "message": "缺少 playwright，请运行 .\\.venv\\Scripts\\python.exe -m pip install playwright",
            "error": str(exc),
        }

    try:
        response = httpx.get(settings.frontend_base_url, timeout=5, trust_env=False)
        result["checks"]["frontend"] = {"ok": response.status_code < 500, "status_code": response.status_code}
        if response.status_code >= 500:
            result["ok"] = False
    except Exception as exc:
        result["ok"] = False
        result["checks"]["frontend"] = {
            "ok": False,
            "message": "前端服务不可访问，请先启动 npm run dev 并确认 FRONTEND_BASE_URL 正确",
            "error": str(exc),
        }
    return result


def write_pdf_report(path: Path, project: SimulationProject, export_task_id: int) -> None:
    preflight = check_pdf_render_prerequisites()
    if not preflight["ok"]:
        failed = [key for key, value in preflight["checks"].items() if isinstance(value, dict) and not value.get("ok")]
        raise RuntimeError(f"pdf_render_failed: PDF 预检失败：{', '.join(failed)}；详情：{json.dumps(preflight, ensure_ascii=False, default=str)}")

    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        raise RuntimeError("pdf_render_failed: 缺少 playwright，请运行 pip install playwright 并安装 chromium") from exc

    token = create_pdf_render_token(project.id, export_task_id)
    url = f"{settings.frontend_base_url.rstrip('/')}/print/{token}"
    try:
        with sync_playwright() as playwright:
            browser = None
            launch_error: Exception | None = None
            for launch_options in (
                {"headless": True},
                {"headless": True, "channel": "msedge"},
                {"headless": True, "channel": "chrome"},
            ):
                try:
                    browser = playwright.chromium.launch(**launch_options)
                    break
                except Exception as exc:
                    launch_error = exc
            if browser is None:
                raise RuntimeError(
                    "Chromium/Edge/Chrome 启动失败；请运行 "
                    ".\\.venv\\Scripts\\python.exe -m playwright install chromium。"
                    f" 原始错误：{launch_error}"
                )
            page = browser.new_page(viewport={"width": 1440, "height": 1200}, device_scale_factor=1)
            page.goto(url, wait_until="networkidle", timeout=45000)
            page.wait_for_function("window.__AGENTSIM_PRINT_READY === true", timeout=45000)
            page.emulate_media(media="print")
            page.evaluate(
                """
                async () => {
                    if (document.fonts && document.fonts.ready) {
                        await document.fonts.ready;
                    }
                    const pendingImages = Array.from(document.images)
                        .filter((image) => !image.complete)
                        .map((image) => new Promise((resolve) => {
                            image.addEventListener("load", resolve, { once: true });
                            image.addEventListener("error", resolve, { once: true });
                        }));
                    await Promise.all(pendingImages);
                    await new Promise((resolve) => requestAnimationFrame(() => requestAnimationFrame(resolve)));
                }
                """
            )
            page.pdf(
                path=str(path),
                format="A4",
                print_background=True,
                prefer_css_page_size=True,
                margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
            )
            browser.close()
    except Exception as exc:
        raise RuntimeError(f"pdf_render_failed: {exc}") from exc


def _append_dict_rows(sheet, rows: list[dict[str, Any]]) -> None:
    rows = _normalize_table_rows(rows)
    keys: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in keys:
                keys.append(key)
    if not keys:
        sheet.append(["暂无数据"])
        return
    sheet.append(keys)
    for row in rows:
        sheet.append([json.dumps(row.get(key), ensure_ascii=False, default=str) if isinstance(row.get(key), (dict, list)) else row.get(key) for key in keys])


def _append_key_value_rows(sheet, data: dict[str, Any]) -> None:
    sheet.append(["字段", "值"])
    if not data:
        sheet.append(["暂无数据", ""])
        return
    for key, value in data.items():
        sheet.append([key, json.dumps(value, ensure_ascii=False, default=str) if isinstance(value, (dict, list)) else value])


def _append_radar_rows(sheet, radar: dict[str, Any]) -> None:
    dimensions = radar.get("dimensions") if isinstance(radar.get("dimensions"), list) else []
    series = radar.get("series") if isinstance(radar.get("series"), list) else []
    if not dimensions or not series:
        sheet.append(["暂无数据"])
        return
    sheet.append(["name", "role", *dimensions])
    for row in series:
        if not isinstance(row, dict):
            continue
        values = row.get("values") if isinstance(row.get("values"), list) else []
        sheet.append([row.get("name"), row.get("role"), *values])


def _create_rows_sheet(workbook, title: str, rows: Any, preferred_empty: str = "暂无数据") -> None:
    sheet = workbook.create_sheet(title[:31])
    _append_dict_rows(sheet, rows)
    if sheet.max_row == 1 and sheet.cell(row=1, column=1).value == "暂无数据":
        sheet.cell(row=1, column=1).value = preferred_empty


def write_excel_report(path: Path, project: SimulationProject) -> None:
    from openpyxl import Workbook

    report = sanitize_report(project.result_data, public=False)
    chart_data = report.get("chart_data") if isinstance(report.get("chart_data"), dict) else {}
    market_config = project.market_config if isinstance(project.market_config, dict) else {}
    workbook = Workbook()
    summary = workbook.active
    summary.title = "报告摘要"
    summary.append(["字段", "值"])
    summary.append(["项目版本", project.plan_type_used or "basic"])
    for key in ("executive_summary", "pricing_analysis", "quality_warnings", "risk_warnings"):
        value = report.get(key)
        summary.append([key, json.dumps(value, ensure_ascii=False, default=str) if isinstance(value, (dict, list)) else value])

    analysis = workbook.create_sheet("业务解读")
    analysis.append(["类型", "内容"])
    for line in build_business_analysis_lines(report):
        analysis.append(["分析", line])
    analysis.append([])
    analysis.append(["维度", "均值", "加权贡献", "权重"])
    for row in _decision_dimension_rows(report):
        analysis.append([row.get("维度"), row.get("均值"), row.get("加权贡献"), row.get("权重")])

    crowd_sheet = workbook.create_sheet("目标人群画像")
    _append_dict_rows(crowd_sheet, crowd_profile_rows(report, market_config))

    selected_strategy_sheet = workbook.create_sheet("已配置策略")
    _append_dict_rows(selected_strategy_sheet, selected_strategy_rows(report, market_config))

    strategy_sheet = workbook.create_sheet("策略建议")
    _append_dict_rows(strategy_sheet, strategy_recommendation_rows(report.get("strategy_recommendations")))

    agents = workbook.create_sheet("Agent样本")
    _append_dict_rows(agents, report.get("agent_samples") if isinstance(report.get("agent_samples"), list) else [])

    decisions = workbook.create_sheet("购买决策")
    _append_dict_rows(decisions, report.get("purchase_decisions") if isinstance(report.get("purchase_decisions"), list) else [])

    aggregation = workbook.create_sheet("聚合指标")
    aggregation.append(["字段", "值"])
    for key, value in (report.get("aggregation") or report.get("metrics") or {}).items():
        aggregation.append([key, json.dumps(value, ensure_ascii=False, default=str) if isinstance(value, (dict, list)) else value])

    evidence = workbook.create_sheet("证据")
    evidence_rows = report.get("evidence_used") or report.get("rag_evidence") or []
    _append_dict_rows(evidence, evidence_rows if isinstance(evidence_rows, list) else [])

    overview = workbook.create_sheet("图表_概览指标")
    _append_key_value_rows(overview, chart_data.get("overview_metrics") if isinstance(chart_data.get("overview_metrics"), dict) else {})

    _create_rows_sheet(workbook, "图表_市场份额", chart_data.get("market_share"))
    _create_rows_sheet(workbook, "图表_全量市场份额", chart_data.get("market_share_full"))
    market_scope = workbook.create_sheet("市场份额口径")
    _append_key_value_rows(market_scope, chart_data.get("market_share_scope") or report.get("market_share_scope") or {})
    _create_rows_sheet(workbook, "市场份额情景", chart_data.get("market_share_scenarios") or report.get("market_share_scenarios"))
    data_gaps = workbook.create_sheet("数据缺口")
    _append_key_value_rows(data_gaps, chart_data.get("data_gaps") or report.get("data_gaps") or {})
    _create_rows_sheet(workbook, "渠道情景", report.get("channel_scenarios"))
    funnel = report.get("propagation_funnel") if isinstance(report.get("propagation_funnel"), dict) else {}
    _create_rows_sheet(workbook, "传播漏斗轮次", funnel.get("rounds"))
    _create_rows_sheet(workbook, "传播漏斗流向", funnel.get("links"))
    _create_rows_sheet(workbook, "舆情演化", funnel.get("sentiment_evolution"))
    _create_rows_sheet(workbook, "竞品逐项分析", chart_data.get("competitor_analysis"))
    _create_rows_sheet(workbook, "图表_购买意愿", chart_data.get("purchase_intent_by_segment"))
    _create_rows_sheet(workbook, "图表_功能重要性", chart_data.get("param_importance"))
    _create_rows_sheet(workbook, "图表_价格敏感", chart_data.get("price_sensitivity"))
    _create_rows_sheet(workbook, "图表_策略ROI", chart_data.get("strategy_roi"))
    _create_rows_sheet(workbook, "图表_渠道贡献", chart_data.get("channel_effect"))
    differentiation = workbook.create_sheet("差异化审计")
    _append_key_value_rows(differentiation, report.get("differentiation_audit") or chart_data.get("differentiation_audit") or {})
    boundaries = workbook.create_sheet("仿真口径")
    _append_key_value_rows(boundaries, report.get("simulation_boundaries") or chart_data.get("simulation_boundaries") or {})
    economics = workbook.create_sheet("策略成本输入")
    _append_key_value_rows(economics, report.get("strategy_economics") or chart_data.get("strategy_economics") or {})
    _create_rows_sheet(workbook, "图表_社交演化", chart_data.get("social_evolution"))
    _create_rows_sheet(workbook, "社交传播轮次", chart_data.get("social_rounds"))

    if isinstance(chart_data.get("competitor_radar"), dict):
        radar = workbook.create_sheet("图表_竞品雷达")
        _append_radar_rows(radar, chart_data["competitor_radar"])
    if isinstance(chart_data.get("competitor_radar_full"), dict):
        radar_full = workbook.create_sheet("图表_全量竞品雷达")
        _append_radar_rows(radar_full, chart_data["competitor_radar_full"])
    if isinstance(chart_data.get("sensitivity_waterfall"), list):
        _create_rows_sheet(workbook, "图表_敏感瀑布", chart_data.get("sensitivity_waterfall"))
    workbook.save(path)


def write_export_file(task: ExportTask, project: SimulationProject) -> Path:
    if task.format not in {"json", "markdown", "excel", "pdf"}:
        raise HTTPException(status_code=422, detail="unsupported_format")

    path = export_file_path(task)
    path.parent.mkdir(parents=True, exist_ok=True)
    if task.format == "json":
        payload = build_report_payload(project, public=False)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    elif task.format == "markdown":
        path.write_text(render_markdown_report(project), encoding="utf-8")
    elif task.format == "pdf":
        write_pdf_report(path, project, task.id)
    else:
        write_excel_report(path, project)

    task.status = "completed"
    task.download_url = f"/api/exports/{task.id}/download"
    task.completed_at = utc_now_naive()
    return path
